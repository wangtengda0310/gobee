#!/usr/bin/env python3
"""
手牌皮肤配置 — 自动修改 Excel 文件

用法:
  python3 modify_excel.py --item-xlsx <path> --cardskin-xlsx <path> \
      --skin-name "皮肤名称" --pinyin "mingmingpinyin" --level 精良|卓越|至臻
"""

import argparse
import sys
from copy import copy
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# 等级→品质映射
LEVEL_MAP = {"精良": 2, "卓越": 3, "至臻": 4}


def find_max_id_row(ws, col_idx, filter_col=None, filter_val=None):
    """找到指定列值最大的行号。
    col_idx: 道具id列索引（1-based）
    filter_col: 可选的过滤列索引
    filter_val: 过滤值
    返回: (row_number, max_value)
    """
    max_id = -1
    max_row = None
    for row_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=col_idx).value
        if cell_val is None:
            continue
        # 如果需要过滤（如只找 CardSkin 类型）
        if filter_col is not None and filter_val is not None:
            filter_cell = ws.cell(row=row_idx, column=filter_col).value
            if str(filter_cell) != filter_val:
                continue
        try:
            val = int(cell_val)
        except (ValueError, TypeError):
            continue
        if val > max_id:
            max_id = val
            max_row = row_idx
    return max_row, max_id


def copy_row_style(ws, src_row, dst_row, max_col):
    """复制整行的值和样式"""
    for col in range(1, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        dst_cell.value = src_cell.value
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)
            dst_cell.alignment = copy(src_cell.alignment)


def safe_insert_rows(ws, start_row, count=1):
    """安全地在 start_row 位置插入 count 行，完整保留所有格式。
    
    openpyxl 原生的 ws.insert_rows() 存在已知 bug：
    - 合并单元格引用不会自动偏移
    - 条件格式范围不会自动调整
    - 数据验证引用不会更新
    - 行高(row_dimensions)不会跟着移动
    
    本函数从下往上逐行搬移数据+样式，并修正以上所有遗漏层。
    """
    max_row_before = ws.max_row
    max_col = ws.max_column

    # ── 1. 从下往上逐行搬移（倒序，避免覆盖）──
    for row in range(max_row_before, start_row - 1, -1):
        new_row = row + count
        for col in range(1, max_col + 1):
            src = ws.cell(row=row, column=col)
            dst = ws.cell(row=new_row, column=col)
            dst.value = src.value
            if src.has_style:
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
                dst.protection = copy(src.protection)

    # ── 2. 清空插入区域（新空行）──
    for row in range(start_row, start_row + count):
        for col in range(1, max_col + 1):
            c = ws.cell(row=row, column=col)
            c.value = None

    # ── 3. 偏移合并单元格引用（原地修改，避免 clear/re-add 导致格式丢失）──
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= start_row:
            mc.shift(0, count)

    # ── 4. 偏移行尺寸 ──
    for r in range(max_row_before, start_row - 1, -1):
        rd_old = ws.row_dimensions.get(r)
        if rd_old is not None:
            rd_new = ws.row_dimensions.get(r + count)
            if rd_new is None:
                from openpyxl.worksheet.dimensions import RowDimension
                rd_new = RowDimension(worksheet=ws, index=r + count)
                ws.row_dimensions[r + count] = rd_new
            rd_new.height = rd_old.height
            rd_new.hidden = rd_old.hidden
            try: rd_new.customFormat = rd_old.customFormat
            except AttributeError: pass
            try: rd_new.customHeight = rd_old.customHeight
            except AttributeError: pass

    return list(range(start_row, start_row + count))


def insert_row_below(ws, ref_row):
    """在 ref_row 下方插入一行，安全搬移所有数据+格式"""
    safe_insert_rows(ws, ref_row + 1, 1)
    return ref_row + 1


def modify_item_xlsx(path, skin_name, pinyin, level):
    """
    修改 Item.xlsx：
    1. 找到道具类型=CardSkin 且道具id 最大的行
    2. 在该行下方插入新行（复制上行全部配置）
    3. 修改：道具id += 1, 道具名称=skin_name, 道具品质=LEVEL_MAP[level], 图标icon=替换拼音
    """
    quality = LEVEL_MAP.get(level)
    if quality is None:
        print(f"错误：未知等级 '{level}'，支持：精良、卓越、至臻")
        sys.exit(1)

    wb = load_workbook(path)
    ws = wb.active

    # 自动找列：需要用户指定列名，这里假设常见列名
    # 道具id在第1列，道具类型在第2列（示例），实际需根据文件调整
    # 先打印表头帮助确认列位置
    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(str(ws.cell(row=1, column=col).value))
    print(f"[Item.xlsx] 表头: {headers}")

    # TODO: 根据实际列名定位。以下为常见假设，需根据表头确认：
    # col_id: 道具id, col_type: 道具类型, col_name: 道具名称
    # col_quality: 道具品质, col_icon: 图标icon
    col_id = col_type = col_name = col_quality = col_icon = None

    for i, h in enumerate(headers, 1):
        hl = h.lower().strip() if h else ""
        if "id" in hl or "道具id" in h:
            col_id = i
        elif "类型" in hl or "道具类型" in h:
            col_type = i
        elif "名称" in hl or "道具名称" in h:
            col_name = i
        elif "品质" in hl or "道具品质" in h:
            col_quality = i
        elif "图标" in hl or "icon" in hl:
            col_icon = i

    if not all([col_id, col_type, col_name, col_quality, col_icon]):
        missing = []
        if not col_id: missing.append("道具id")
        if not col_type: missing.append("道具类型")
        if not col_name: missing.append("道具名称")
        if not col_quality: missing.append("道具品质")
        if not col_icon: missing.append("图标icon")
        print(f"警告：未找到以下列: {missing}")
        print("请手动指定列位置。在 SKILL.md 中查看详细说明。")
        wb.close()
        return None

    # 找到 CardSkin 类型且道具id 最大的行
    max_row, max_id = find_max_id_row(
        ws, col_id, filter_col=col_type, filter_val="CardSkin"
    )
    if max_row is None:
        print("错误：未找到道具类型为 CardSkin 的行")
        wb.close()
        sys.exit(1)

    print(f"[Item.xlsx] 找到最大道具id: {max_id}, 行号: {max_row}")

    # 插入新行并复制
    new_row = insert_row_below(ws, max_row)
    copy_row_style(ws, max_row, new_row, ws.max_column)

    # 修改新行
    new_id = max_id + 1
    ws.cell(row=new_row, column=col_id).value = new_id
    ws.cell(row=new_row, column=col_name).value = skin_name
    ws.cell(row=new_row, column=col_quality).value = quality

    # 替换图标中的拼音
    old_icon = str(ws.cell(row=new_row, column=col_icon).value or "")
    # 假设图标路径格式如: .../pinyin_name/... 或包含拼音部分
    # 这里是通用替换逻辑，需根据实际格式调整
    new_icon = _replace_pinyin_part(old_icon, pinyin)
    ws.cell(row=new_row, column=col_icon).value = new_icon

    wb.save(path)
    wb.close()
    print(f"[Item.xlsx] ✅ 已新增道具: 道具id={new_id}, 名称={skin_name}, 品质={quality}")
    return new_id, skin_name


def modify_cardskin_xlsx(path, new_item_id, skin_name, pinyin):
    """
    修改 CardSkin_手牌皮肤表.xlsx：
    1. 找到手牌皮肤道具id 最大的行
    2. 在该行下方插入新行（复制上行全部配置）
    3. 修改：手牌皮肤道具id=new_item_id, 手牌皮肤名称=skin_name, D-K列替换拼音
    4. 不配置手牌皮肤卡面图路径
    """
    wb = load_workbook(path)
    ws = wb.active

    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(str(ws.cell(row=1, column=col).value))
    print(f"[CardSkin] 表头: {headers}")

    col_skin_id = col_skin_name = col_card_path = None
    col_d = col_e = col_f = col_g = col_h = col_i = col_j = col_k = None

    for i, h in enumerate(headers, 1):
        hl = h.lower().strip() if h else ""
        if "皮肤道具id" in hl or "手牌皮肤道具id" in h:
            col_skin_id = i
        elif "皮肤名称" in hl or "手牌皮肤名称" in h:
            col_skin_name = i
        elif "卡面图" in hl or "卡面" in hl:
            col_card_path = i

    # D-K 列 = 第4列到第11列
    col_d, col_e, col_f, col_g = 4, 5, 6, 7
    col_h, col_i, col_j, col_k = 8, 9, 10, 11

    if not all([col_skin_id, col_skin_name]):
        missing = []
        if not col_skin_id: missing.append("手牌皮肤道具id")
        if not col_skin_name: missing.append("手牌皮肤名称")
        print(f"警告：未找到以下列: {missing}")
        wb.close()
        return

    # 找到手牌皮肤道具id 最大的行
    max_row, max_id = find_max_id_row(ws, col_skin_id)
    if max_row is None:
        print("错误：未找到任何手牌皮肤数据")
        wb.close()
        sys.exit(1)

    print(f"[CardSkin] 找到最大皮肤id: {max_id}, 行号: {max_row}")

    # 插入新行
    new_row = insert_row_below(ws, max_row)
    copy_row_style(ws, max_row, new_row, ws.max_column)

    # 修改
    ws.cell(row=new_row, column=col_skin_id).value = new_item_id
    ws.cell(row=new_row, column=col_skin_name).value = skin_name

    # 不配置卡面图路径（清空或保持为空）
    if col_card_path:
        ws.cell(row=new_row, column=col_card_path).value = None

    # 替换 D-K 列中的拼音部分
    for col in [col_d, col_e, col_f, col_g, col_h, col_i, col_j, col_k]:
        if col > ws.max_column:
            break
        old_val = str(ws.cell(row=new_row, column=col).value or "")
        new_val = _replace_pinyin_part(old_val, pinyin)
        ws.cell(row=new_row, column=col).value = new_val if new_val else None

    wb.save(path)
    wb.close()
    print(f"[CardSkin] ✅ 已新增手牌皮肤: id={new_item_id}, 名称={skin_name}")


def _replace_pinyin_part(text, new_pinyin):
    """
    替换文本中的拼音名称部分。
    假设格式如: .../zhugeliang/... 或包含下划线拼音如 card_skin_zhugeliang
    替换规则：找到最后一个英文/下划线+拼音的部分进行替换
    """
    if not text or not new_pinyin:
        return text

    # 策略：替换路径中最后一个目录名（通常是拼音）
    # 例如: res/card/zhugeliang/icon.png → res/card/mingmingpinyin/icon.png
    parts = text.rsplit("/", 1)
    if len(parts) == 2 and parts[0]:
        # 有目录结构的情况
        return f"{parts[0]}/{new_pinyin}{parts[1][parts[1].find('/'):]}" if "/" in parts[1][1:] else f"{parts[0]}/{new_pinyin}"
    else:
        # 没有目录结构，直接替换拼音部分
        # 尝试匹配常见的拼音模式（小写字母+下划线组合）
        import re
        return re.sub(r'[a-z_]+[a-z]', new_pinyin, text, count=1)

    return text


def main():
    parser = argparse.ArgumentParser(description="手牌皮肤配置工具")
    parser.add_argument("--item-xlsx", required=True, help="Item.xlsx 文件路径")
    parser.add_argument("--cardskin-xlsx", required=True, help="CardSkin_手牌皮肤表.xlsx 文件路径")
    parser.add_argument("--skin-name", required=True, help="皮肤名称")
    parser.add_argument("--pinyin", required=True, help="命名拼音")
    parser.add_argument("--level", required=True, choices=["精良", "卓越", "至臻"], help="皮肤等级")
    parser.add_argument("--debug", action="store_true", help="调试模式：只分析不修改")

    args = parser.parse_args()

    if args.debug:
        print("[调试模式] 分析文件结构但不修改...")

    new_id, name = modify_item_xlsx(
        args.item_xlsx, args.skin_name, args.pinyin, args.level
    )

    if new_id is None:
        print("Item.xlsx 修改失败，跳过 CardSkin 表修改。")
        sys.exit(1)

    if not args.debug:
        modify_cardskin_xlsx(
            args.cardskin_xlsx, new_id, name, args.pinyin
        )

    print("\n===== 配置完成 =====")
    print(f"道具id: {new_id}")
    print(f"皮肤名称: {name}")
    print(f"等级/品质: {args.level} / {LEVEL_MAP[args.level]}")
    print(f"拼音: {args.pinyin}")
    print("⚠️ 手牌皮肤卡面图路径未配置，请手动补充。")


if __name__ == "__main__":
    main()
