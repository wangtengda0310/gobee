#!/usr/bin/env python3
"""生成名将杀技能测试用例 Excel + Markdown 双份交付。

用法:
    python generate_outputs.py <overview.json> <testcases.json> <output_dir> <skill_name>

将在 output_dir 下生成:
    - <skill_name>_测试用例.xlsx
    - <skill_name>_测试用例.md
"""

import json
import sys
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════
# 样式常量
# ═══════════════════════════════════════════════════════════

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=12, color="1F4E79")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TITLE_FONT = Font(bold=True, size=16)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def apply_border(ws, row, col):
    ws.cell(row=row, column=col).border = THIN_BORDER


# ═══════════════════════════════════════════════════════════
# Sheet 1: 测试概览
# ═══════════════════════════════════════════════════════════

def write_overview_sheet(ws, data):
    row = 1
    title = data.get("title", "名将杀技能测试概览")
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    row += 2

    info_pairs = [
        ("技能名称", data.get("skill_name", "")),
        ("所属武将", data.get("hero", "")),
    ]
    if data.get("faction"):
        info_pairs.append(("武将势力", data.get("faction")))
    if data.get("skill_category"):
        info_pairs.append(("技能分类", data.get("skill_category")))
    for label, value in info_pairs:
        ws.cell(row=row, column=1, value=label).font = HEADER_FONT
        cell = ws.cell(row=row, column=2, value=str(value))
        cell.alignment = WRAP_ALIGN
        row += 1

    # 技能条目（entries）；兼容旧 skill_text
    entries = data.get("entries", [])
    if entries:
        for ent in entries:
            label = f"条目 {ent.get('entry_id', '')}（{ent.get('发动方式', '')}）"
            ws.cell(row=row, column=1, value=label).font = HEADER_FONT
            cell = ws.cell(row=row, column=2, value=str(ent.get("文案", "")))
            cell.alignment = WRAP_ALIGN
            row += 1
    elif data.get("skill_text"):
        ws.cell(row=row, column=1, value="技能文案（原文）").font = HEADER_FONT
        cell = ws.cell(row=row, column=2, value=str(data.get("skill_text", "")))
        cell.alignment = WRAP_ALIGN
        row += 1

    subs = data.get("sub_skills", [])
    if subs:
        row += 1
        for si, sub in enumerate(subs):
            for label, value in sub.items():
                ws.cell(row=row, column=1, value=label).font = HEADER_FONT
                cell = ws.cell(row=row, column=2, value=str(value))
                cell.alignment = WRAP_ALIGN
                row += 1
            if si < len(subs) - 1:
                row += 1

    sections = data.get("sections", [])
    for sec in sections:
        row += 1
        heading = sec.get("heading", "")
        cell = ws.cell(row=row, column=1, value=heading)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        col_count = len(sec.get("columns", [])) or 1
        if col_count > 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
            for c in range(2, col_count + 1):
                ws.cell(row=row, column=c).fill = SECTION_FILL
        row += 1

        for ci, col_name in enumerate(sec.get("columns", []), 1):
            cell = ws.cell(row=row, column=ci, value=col_name)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        row += 1

        for r_data in sec.get("rows", []):
            for ci, val in enumerate(r_data, 1):
                cell = ws.cell(row=row, column=ci, value=str(val) if val else "")
                cell.alignment = WRAP_ALIGN
                apply_border(ws, row, ci)
            row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 75
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 45


# ═══════════════════════════════════════════════════════════
# Sheet 2: 测试用例
# ═══════════════════════════════════════════════════════════

def write_testcases_sheet(ws, data):
    columns = data.get(
        "columns",
        ["编号", "一级模块", "二级模块", "拆解项", "标题", "前置条件", "步骤", "预期结果"],
    )
    rows = data.get("rows", [])

    for ci, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for ri, row_data in enumerate(rows, 2):
        if isinstance(row_data, dict):
            vals = [row_data.get(col, "") for col in columns]
        else:
            vals = row_data
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=str(val) if val else "")
            cell.alignment = WRAP_ALIGN
            apply_border(ws, ri, ci)

    col_widths = [8, 14, 16, 10, 30, 32, 34, 50]
    for i, w in enumerate(col_widths[: len(columns)], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{1 + len(rows)}"
    ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════
# Markdown 生成
# ═══════════════════════════════════════════════════════════

def _md_escape(text):
    """转义 Markdown 特殊字符（表格内换行用 <br>）。"""
    if not text:
        return ""
    text = str(text)
    # 表格内换行
    text = text.replace("\n", "<br>")
    # 转义管道符
    text = text.replace("|", "\\|")
    return text


def _md_table(columns, rows):
    """生成 Markdown 表格字符串。"""
    lines = []
    lines.append("| " + " | ".join(_md_escape(c) for c in columns) + " |")
    lines.append("| " + " | ".join("---" for _ in columns) + " |")
    for row in rows:
        if isinstance(row, dict):
            row_vals = [row.get(col, "") for col in columns]
        else:
            row_vals = row
        cells = [_md_escape(cell) for cell in row_vals]
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def generate_markdown(overview, testcases):
    """从 overview + testcases JSON 生成完整 Markdown 报告。"""
    lines = []

    # 标题
    title = overview.get("title", "名将杀技能测试概览")
    skill_name = overview.get("skill_name", "")
    lines.append(f"# {title}")
    lines.append("")

    # 基本信息
    lines.append(f"**技能名称**：{_md_escape(skill_name)}")
    lines.append(f"**所属武将**：{_md_escape(overview.get('hero', ''))}")
    if overview.get("faction"):
        lines.append(f"**武将势力**：{_md_escape(overview.get('faction'))}")
    if overview.get("skill_category"):
        lines.append(f"**技能分类**：{_md_escape(overview.get('skill_category'))}")
    lines.append("")

    # 技能条目（entries）；兼容旧 skill_text
    entries = overview.get("entries", [])
    if entries:
        lines.append("**技能条目（原文）**：")
        for ent in entries:
            label = f"条目 {ent.get('entry_id', '')}（{ent.get('发动方式', '')}）"
            lines.append(f"- {_md_escape(label)}：{_md_escape(ent.get('文案', ''))}")
        lines.append("")
    elif overview.get("skill_text"):
        lines.append(f"**技能文案（原文）**：{_md_escape(overview.get('skill_text', ''))}")
        lines.append("")

    # 子技能
    subs = overview.get("sub_skills", [])
    if subs:
        lines.append("## 子技能")
        lines.append("")
        for sub in subs:
            for label, value in sub.items():
                lines.append(f"- **{label}**：{_md_escape(value)}")
            lines.append("")
        lines.append("")

    # 各小节
    sections = overview.get("sections", [])
    for sec in sections:
        heading = sec.get("heading", "")
        lines.append(f"## {heading}")
        lines.append("")
        columns = sec.get("columns", [])
        rows = sec.get("rows", [])
        if columns and rows:
            lines.append(_md_table(columns, rows))
        elif rows:
            for row in rows:
                lines.append("- " + " / ".join(_md_escape(c) for c in row))
        lines.append("")

    # 测试用例
    tc_columns = testcases.get(
        "columns",
        ["编号", "一级模块", "二级模块", "拆解项", "标题", "前置条件", "步骤", "预期结果"],
    )
    tc_rows = testcases.get("rows", [])
    lines.append(f"## 测试用例（共 {len(tc_rows)} 条）")
    lines.append("")
    if tc_rows:
        lines.append(_md_table(tc_columns, tc_rows))
        lines.append("")

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════

def main():
    if len(sys.argv) != 5:
        print(f"用法: {sys.argv[0]} <overview.json> <testcases.json> <output_dir> <skill_name>")
        sys.exit(1)

    overview_path = sys.argv[1]
    testcases_path = sys.argv[2]
    output_dir = sys.argv[3]
    skill_name = sys.argv[4]

    with open(overview_path, "r", encoding="utf-8") as f:
        overview = json.load(f)
    with open(testcases_path, "r", encoding="utf-8") as f:
        testcases = json.load(f)

    os.makedirs(output_dir, exist_ok=True)

    # ── Excel ──
    xlsx_path = os.path.join(output_dir, f"{skill_name}_测试用例.xlsx")
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "测试概览"
    write_overview_sheet(ws1, overview)

    ws2 = wb.create_sheet("测试用例")
    write_testcases_sheet(ws2, testcases)

    wb.save(xlsx_path)
    print(f"✅ Excel: {xlsx_path}")

    # ── Markdown ──
    md_path = os.path.join(output_dir, f"{skill_name}_测试用例.md")
    md_content = generate_markdown(overview, testcases)
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md_content)
    print(f"✅ Markdown: {md_path}")


if __name__ == "__main__":
    main()
