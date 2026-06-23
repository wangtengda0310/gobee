#!/usr/bin/env python3
"""
形象边框配置 — 在 Item.xlsx 和 ItemFrame_边框道具表.xlsx 中插入新边框行（保留样式）

使用 openpyxl 操作 Excel，完整保留单元格字体、边框、填充、对齐等样式。

用法:
  python3 configure_frame.py <base_dir> <frame_name> <pinyin_name> <level>

参数:
  base_dir   - Excel 文件所在目录 (如 /path/to/名将杀配置/)
  frame_name - 形象边框名称 (如 "棠花沁梦")
  pinyin     - 命名拼音 (如 "tanghuaqinmeng")
  level      - "精良" | "卓越" | "至臻"

示例:
  python3 configure_frame.py "/path/to/名将杀配置/" "棠花沁梦" "tanghuaqinmeng" "至臻"
"""

import sys
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 等级→品质映射
LEVEL_MAP = {"精良": 2, "卓越": 3, "至臻": 4}


def copy_row_style(ws, src_row, dst_row, max_col):
    """复制整行的值和样式（字体/边框/填充/对齐/数字格式/保护）"""
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        dst.value = src.value
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)


def safe_insert_rows(ws, start_row, count=1):
    """安全地在 start_row 位置插入 count 行，完整保留所有格式。
    
    openpyxl 原生的 ws.insert_rows() 存在已知 bug：
    - 合并单元格引用不会自动偏移
    - 条件格式范围不会自动调整
    - 数据验证引用不会更新
    - 行高(row_dimensions)不会跟着移动
    
    本函数从下往上逐行搬移数据+样式，并修正以上所有遗漏层。
    """
    from openpyxl.worksheet.dimensions import RowDimension

    max_row_before = ws.max_row
    max_col = ws.max_column

    # ── 0. 展开行级样式到单元格（解决 openpyxl 无法搬移 row.s 的问题）──
    # 仅处理分节标题行（第一列为 '#' 开头），避免污染数据行
    for row in range(start_row, max_row_before + 1):
        first_val = ws.cell(row=row, column=1).value
        if not first_val or not str(first_val).startswith('#'):
            continue
        rd = ws.row_dimensions.get(row)
        if rd is not None and getattr(rd, 's', None):
            default_fill = None
            for c in range(1, max_col + 1):
                cell = ws.cell(row=row, column=c)
                f = cell.fill
                if f and f.fill_type and f.start_color and f.start_color.rgb and f.start_color.rgb != '00000000':
                    default_fill = copy(f)
                    break
            if default_fill:
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=c)
                    f = cell.fill
                    if not f or not f.fill_type or not f.start_color or f.start_color.rgb == '00000000':
                        cell.fill = copy(default_fill)

    # ── 1. 从下往上逐行搬移（倒序，避免覆盖）──
    for row in range(max_row_before, start_row - 1, -1):
        new_row = row + count
        for col in range(1, max_col + 1):
            src = ws.cell(row=row, column=col)
            dst = ws.cell(row=new_row, column=col)
            dst.value = src.value
            if src.has_style:
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
                dst.protection = copy(src.protection)
            else:
                # 重置目标格样式，避免残留上一轮搬移的旧样式
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
                dst.protection = copy(src.protection)

    # ── 2. 清空插入区域 ──
    for row in range(start_row, start_row + count):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None

    # ── 3. 偏移合并单元格引用（原地修改，避免 clear/re-add 导致格式丢失）──
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= start_row:
            mc.shift(0, count)

    # ── 4. 偏移行尺寸 ──
    for r in range(max_row_before, start_row - 1, -1):
        rd_old = ws.row_dimensions.get(r)
        if rd_old is not None:
            rd_new = ws.row_dimensions.get(r + count)
            if rd_new is None:
                rd_new = RowDimension(worksheet=ws, index=r + count)
                ws.row_dimensions[r + count] = rd_new
            rd_new.height = rd_old.height
            rd_new.hidden = rd_old.hidden
            try: rd_new.customFormat = rd_old.customFormat
            except AttributeError: pass
            try: rd_new.customHeight = rd_old.customHeight
            except AttributeError: pass
            if hasattr(rd_old, 's'):
                try: rd_new.s = rd_old.s
                except AttributeError: pass

    # ── 5. 清除非标题行的行级样式（s 只读无法搬移，删除 RowDimension 消除）──
    # 包括插入区域，新行不应继承旧行的 row.s
    for row in range(start_row, max_row_before + 1):
        rd = ws.row_dimensions.get(row)
        if rd and getattr(rd, 's', None):
            first_val = ws.cell(row=row, column=1).value
            if not first_val or not str(first_val).startswith('#'):
                h = rd.height
                hid = rd.hidden
                del ws.row_dimensions[row]
                if h is not None or hid:
                    nr = RowDimension(worksheet=ws, index=row)
                    nr.height = h
                    nr.hidden = hid
                    ws.row_dimensions[row] = nr

    return list(range(start_row, start_row + count))


def insert_row_below(ws, ref_row):
    """在 ref_row 下方插入一行，安全搬移所有数据+格式"""
    safe_insert_rows(ws, ref_row + 1, 1)
    return ref_row + 1


def find_max_id_row(ws, col_idx, filter_col=None, filter_val=None, data_start_row=5):
    """找到 col_idx 列中值最大的行号（从 data_start_row 开始，过滤表头行）。

    返回: (row_number, max_value)
    """
    max_id = -1
    max_row = None
    for row_idx in range(data_start_row, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=col_idx).value
        if cell_val is None:
            continue
        # 过滤列条件
        if filter_col is not None and filter_val is not None:
            filter_cell = ws.cell(row=row_idx, column=filter_col).value
            if str(filter_cell) != filter_val:
                continue
        try:
            val = int(cell_val)
        except (ValueError, TypeError):
            continue
        if val > max_id:
            max_id = val
            max_row = row_idx
    return max_row, max_id


def find_existing_row(ws, col_idx, search_value, filter_col=None, filter_val=None, data_start_row=5):
    """查找指定列中值匹配的行号。

    返回: 行号 (int) 或 None（未找到）
    """
    for row_idx in range(data_start_row, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=col_idx).value
        if cell_val is not None and str(cell_val).strip() == str(search_value).strip():
            # 如果指定了过滤条件
            if filter_col is not None and filter_val is not None:
                filter_cell = ws.cell(row=row_idx, column=filter_col).value
                if str(filter_cell) == str(filter_val):
                    return row_idx
            else:
                return row_idx
    return None


def extract_pinyin(cell_c_val):
    """从边框表 C 列标识中提取旧拼音。

    "ui_s1_frame_character_songfengheying_01" → "songfengheying"
    "ui_s1_player_kuang_jinque" → "jinque"
    """
    if not cell_c_val:
        return ""
    s = str(cell_c_val)
    s = re.sub(r'^ui_s1_(player_kuang|frame_character|frame)_', '', s)
    s = re.sub(r'_\d+$', '', s)
    return s


def extract_pascal_dir(cell_val):
    """从 Prefab 路径中提取 PascalCase 目录名。

    "Prefabs/.../FrameSkin/SongFengHeYing/fx_ui_..." → "SongFengHeYing"
    """
    if not cell_val:
        return ""
    segs = str(cell_val).split('/')
    for seg in reversed(segs):
        if re.search(r'[A-Z]', seg) and not seg.startswith('fx_'):
            return seg
    return ""


def replace_pinyin_in_text(text, old_pinyin, new_pinyin, old_pascal="", new_pascal=""):
    """替换文本中的旧拼音→新拼音，旧PascalCase→新PascalCase"""
    if not text:
        return text
    result = str(text)
    if old_pinyin:
        result = result.replace(old_pinyin, new_pinyin)
    if old_pascal and new_pascal:
        result = result.replace(old_pascal, new_pascal)
    return result


def pinyin_to_pascalcase(pinyin):
    """将拼音转为按音节分割的 PascalCase。

    "tanghuaqinmeng" → "TangHuaQinMeng"
    """
    initials = r'(?:zh|ch|sh|[bpmfdtnlgkhjqxrzcsyw])'
    # 在声母前插入分隔符（要求声母后跟元音字母，避免误拆韵尾 n/ng）
    marked = re.sub(r'(' + initials + r')(?=[aeiouv])', r'|\1', pinyin)
    syllables = [s for s in marked.split('|') if s]
    if not syllables:
        return pinyin.capitalize()
    return ''.join(s.capitalize() for s in syllables)


# ── 第一步: Item.xlsx ──────────────────────────────────────

def modify_item_xlsx(path, frame_name, pinyin, level):
    quality = LEVEL_MAP.get(level)
    if quality is None:
        print(f"错误: 未知等级 '{level}'，支持: 精良、卓越、至臻")
        sys.exit(1)

    wb = load_workbook(path)
    ws = wb.active  # 道具表|Item

    # 列位置（硬编码，基于已知表结构）
    # A=1: 道具id  B=2: 道具名称  C=3: 道具类型  D=4: 道具品质  V=22: 图标icon
    COL_ID = 1
    COL_TYPE = 3
    COL_NAME = 2
    COL_QUALITY = 4
    COL_ICON = 22

    # ── 先检查是否已存在同名 Frame 道具 ──
    existing_row = find_existing_row(ws, COL_NAME, frame_name,
                                     filter_col=COL_TYPE, filter_val="Frame")
    if existing_row is not None:
        existing_id = ws.cell(row=existing_row, column=COL_ID).value
        modified = False

        # 校验并修正品质
        cur_quality = ws.cell(row=existing_row, column=COL_QUALITY).value
        if cur_quality != quality:
            ws.cell(row=existing_row, column=COL_QUALITY).value = quality
            print(f"[Item.xlsx] 🔧 行 {existing_row} 道具品质: {cur_quality} → {quality}")
            modified = True

        # 校验并修正图标拼音
        cur_icon = str(ws.cell(row=existing_row, column=COL_ICON).value or "")
        new_icon = re.sub(
            r'ui_s1_daoju_frame_character_[a-z0-9_]+(?=\.)',
            f'ui_s1_daoju_frame_character_{pinyin}',
            cur_icon
        )
        if cur_icon != new_icon:
            ws.cell(row=existing_row, column=COL_ICON).value = new_icon
            print(f"[Item.xlsx] 🔧 行 {existing_row} 图标: {cur_icon} → {new_icon}")
            modified = True

        if modified:
            wb.save(path)
            print(f"[Item.xlsx] ✅ 已修正已有 Frame: ID={existing_id}, Name={frame_name}")
        else:
            print(f"[Item.xlsx] ⏭️ Frame 已存在且正确: ID={existing_id}, Name={frame_name}, 跳过")
        wb.close()
        return existing_id

    # ── 不存在则新增 ──
    max_row, max_id = find_max_id_row(ws, COL_ID, filter_col=COL_TYPE, filter_val="Frame")
    if max_row is None:
        print("错误: 未找到道具类型为 Frame 的行")
        wb.close()
        sys.exit(1)

    print(f"[Item.xlsx] 最大 Frame ID: {max_id} (行 {max_row})")
    new_id = max_id + 1

    # 插入新行
    new_row = insert_row_below(ws, max_row)
    copy_row_style(ws, max_row, new_row, ws.max_column)

    # 修改新行
    ws.cell(row=new_row, column=COL_ID).value = new_id
    ws.cell(row=new_row, column=COL_NAME).value = frame_name
    ws.cell(row=new_row, column=COL_QUALITY).value = quality

    # 图标
    old_icon = str(ws.cell(row=new_row, column=COL_ICON).value or "")
    new_icon = re.sub(
        r'ui_s1_daoju_frame_character_[a-z0-9_]+(?=\.)',
        f'ui_s1_daoju_frame_character_{pinyin}',
        old_icon
    )
    ws.cell(row=new_row, column=COL_ICON).value = new_icon

    wb.save(path)
    wb.close()
    print(f"[Item.xlsx] ✅ 新 Frame: ID={new_id}, Name={frame_name}, Rarity={quality}")
    print(f"             Icon: {new_icon}")
    return new_id


# ── 第二步: ItemFrame_边框道具表.xlsx ─────────────────────

def modify_frame_xlsx(path, new_item_id, frame_name, pinyin):
    wb = load_workbook(path)
    ws = wb.active  # 边框表|FrameItem

    # 列位置
    # A=1: 道具id  B=2: 边框名  C=3: 标识  D=4: Image路径
    # E=5: 头像边框  F=6: Prefab路径  G=7: Head边框路径
    COL_ID = 1
    COL_NAME = 2
    COL_C = 3
    COL_D = 4
    COL_E = 5
    COL_F = 6
    COL_G = 7

    # ── 先检查是否已存在同名边框 ──
    existing_row = find_existing_row(ws, COL_NAME, frame_name)
    if existing_row is not None:
        existing_id = ws.cell(row=existing_row, column=COL_ID).value
        modified = False

        # 从已有行提取旧拼音和旧目录，用于对比
        old_pinyin = extract_pinyin(ws.cell(row=existing_row, column=COL_C).value)
        old_pascal = extract_pascal_dir(
            ws.cell(row=existing_row, column=COL_F).value or
            ws.cell(row=existing_row, column=COL_G).value
        )
        new_pascal = pinyin_to_pascalcase(pinyin)

        # 校验道具 ID
        cur_id = ws.cell(row=existing_row, column=COL_ID).value
        if cur_id != new_item_id:
            ws.cell(row=existing_row, column=COL_ID).value = new_item_id
            print(f"[ItemFrame] 🔧 行 {existing_row} 道具id: {cur_id} → {new_item_id}")
            modified = True

        # 校验 C~G 列拼音
        checks = [
            (COL_C, "C(标识)"), (COL_D, "D(图片)"), (COL_E, "E(头像)"),
            (COL_F, "F(Prefab)"), (COL_G, "G(Head)"),
        ]
        for col, label in checks:
            cur_val = str(ws.cell(row=existing_row, column=col).value or "")
            if col in (COL_F, COL_G):
                expected = replace_pinyin_in_text(cur_val, old_pinyin, pinyin, old_pascal, new_pascal)
            else:
                expected = replace_pinyin_in_text(cur_val, old_pinyin, pinyin)
            if cur_val != expected:
                ws.cell(row=existing_row, column=col).value = expected if expected else None
                print(f"[ItemFrame] 🔧 行 {existing_row} {label}: {cur_val} → {expected}")
                modified = True

        if modified:
            wb.save(path)
            print(f"[ItemFrame] ✅ 已修正已有边框: ID={existing_id}, Name={frame_name}")
        else:
            print(f"[ItemFrame] ⏭️ 边框已存在且正确: ID={existing_id}, Name={frame_name}, 跳过")

        # 打印结果
        for col, label in checks:
            print(f"           {label}: {ws.cell(row=existing_row, column=col).value}")
        wb.close()
        return

    # ── 不存在则新增 ──
    max_row, max_id = find_max_id_row(ws, COL_ID)
    if max_row is None:
        print("错误: 未找到边框数据行")
        wb.close()
        sys.exit(1)

    # 提取旧拼音（从 C 列）
    old_pinyin = extract_pinyin(ws.cell(row=max_row, column=COL_C).value)
    # 提取旧 PascalCase 目录名（从 F 或 G 列）
    old_pascal = extract_pascal_dir(
        ws.cell(row=max_row, column=COL_F).value or
        ws.cell(row=max_row, column=COL_G).value
    )
    new_pascal = pinyin_to_pascalcase(pinyin)

    print(f"[ItemFrame] 最大边框 ID: {max_id} (行 {max_row})")
    print(f"           旧拼音: {old_pinyin} → {pinyin}")
    if old_pascal:
        print(f"           旧目录: {old_pascal} → {new_pascal}")

    # 插入新行并复制
    new_row = insert_row_below(ws, max_row)
    copy_row_style(ws, max_row, new_row, ws.max_column)

    # 修改
    ws.cell(row=new_row, column=COL_ID).value = new_item_id
    ws.cell(row=new_row, column=COL_NAME).value = frame_name

    # C~E 列: 仅替换拼音（全小写），不替换 PascalCase 目录名
    for col in [COL_C, COL_D, COL_E]:
        old_val = str(ws.cell(row=new_row, column=col).value or "")
        new_val = replace_pinyin_in_text(old_val, old_pinyin, pinyin)
        ws.cell(row=new_row, column=col).value = new_val if new_val else None

    # F~G 列: 替换拼音 + PascalCase 目录名（音节首字母大写）
    for col in [COL_F, COL_G]:
        old_val = str(ws.cell(row=new_row, column=col).value or "")
        new_val = replace_pinyin_in_text(old_val, old_pinyin, pinyin, old_pascal, new_pascal)
        ws.cell(row=new_row, column=col).value = new_val if new_val else None

    wb.save(path)
    wb.close()

    # 打印结果
    for col, label in [(COL_C, "C(标识)"), (COL_D, "D(图片)"), (COL_E, "E(头像)"),
                        (COL_F, "F(Prefab)"), (COL_G, "G(Head)")]:
        print(f"           {label}: {ws.cell(row=new_row, column=col).value}")


# ── 主入口 ─────────────────────────────────────────────────

def main():
    if len(sys.argv) < 5:
        print("用法: python3 configure_frame.py <base_dir> <frame_name> <pinyin> <level>")
        print("示例: python3 configure_frame.py /path/to/名将杀配置/ 棠花沁梦 tanghuaqinmeng 至臻")
        sys.exit(1)

    base_dir = Path(sys.argv[1])
    frame_name = sys.argv[2]
    pinyin = sys.argv[3]
    level = sys.argv[4]

    item_path = base_dir / "Item.xlsx"
    frame_path = base_dir / "ItemFrame_边框道具表.xlsx"

    level_str = {2: "精良", 3: "卓越", 4: "至臻"}.get(LEVEL_MAP.get(level), level)

    print(f"\n=== 配置形象边框: {frame_name} / {pinyin} / {level_str} ===\n")
    print("▶ 第一步: Item.xlsx")

    new_id = modify_item_xlsx(str(item_path), frame_name, pinyin, level)

    print("\n▶ 第二步: ItemFrame_边框道具表.xlsx")
    modify_frame_xlsx(str(frame_path), new_id, frame_name, pinyin)

    print(f"\n=== ✅ 配置完成! 道具 ID: {new_id} ===\n")


if __name__ == "__main__":
    main()
