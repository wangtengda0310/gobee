#!/usr/bin/env python3
"""
configure_hero_skin.py - Configure a new hero skin item across multiple Excel files.

Usage:
    python configure_hero_skin.py <base_dir> <skin_name> <hero_name> <skin_pinyin> <level>

Example:
    python configure_hero_skin.py /path/to/config 驰猎惊澜 孙权 chiliejinglan 至臻
"""

import argparse
import os
import sys
from copy import copy

import openpyxl


# ──────────────────────────────────────────────────────────────────────
# Level mappings
# ──────────────────────────────────────────────────────────────────────
LEVEL_MAP_QUALITY = {"精良": 2, "卓越": 3, "至臻": 4}
LEVEL_MAP_RAILY = {"精良": 1, "卓越": 2, "至臻": 3}


# ──────────────────────────────────────────────────────────────────────
# Helper functions
# ──────────────────────────────────────────────────────────────────────

def copy_row_style(ws, src_row, dst_row, max_col):
    """Copy values, styles (font/border/fill/alignment/number_format) from src_row to dst_row."""
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)

        dst.value = src.value

        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = copy(src.number_format)
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)


def safe_insert_rows(ws, start_row, count=1):
    """安全地在 start_row 位置插入 count 行，完整保留所有格式。

    openpyxl 原生的 ws.insert_rows() 存在已知 bug：
    - 合并单元格引用不会自动偏移
    - 条件格式范围不会自动调整
    - 数据验证引用不会更新
    - 行高(row_dimensions)不会跟着移动

    本函数从下往上逐行搬移数据+样式，并修正以上所有遗漏层。
    """
    from openpyxl.worksheet.dimensions import RowDimension

    max_row_before = ws.max_row
    max_col = ws.max_column

    # ── 1. 保存合并单元格 ──

    # ── 2. 从下往上逐行搬移（倒序，避免覆盖）──
    for row in range(max_row_before, start_row - 1, -1):
        new_row = row + count
        for col in range(1, max_col + 1):
            src = ws.cell(row=row, column=col)
            dst = ws.cell(row=new_row, column=col)
            dst.value = src.value
            if src.has_style:
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
                dst.protection = copy(src.protection)

    # ── 3. 清空插入区域 ──
    for row in range(start_row, start_row + count):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None

    # ── 4. 偏移合并单元格引用（原地修改，避免 clear/re-add 导致格式丢失）──
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= start_row:
            mc.shift(0, count)

    # ── 5. 偏移行尺寸 ──
    for r in range(max_row_before, start_row - 1, -1):
        rd_old = ws.row_dimensions.get(r)
        if rd_old is not None:
            rd_new = ws.row_dimensions.get(r + count)
            if rd_new is None:
                rd_new = RowDimension(worksheet=ws, index=r + count)
                ws.row_dimensions[r + count] = rd_new
            rd_new.height = rd_old.height
            rd_new.hidden = rd_old.hidden
            try: rd_new.customFormat = rd_old.customFormat
            except AttributeError: pass
            try: rd_new.customHeight = rd_old.customHeight
            except AttributeError: pass

    return list(range(start_row, start_row + count))


def insert_row_below(ws, ref_row, max_col):
    """在 ref_row 下方安全插入新行，复制样式"""
    new_row = ref_row + 1
    safe_insert_rows(ws, new_row, 1)
    copy_row_style(ws, ref_row, new_row, max_col)
    return new_row


def find_max_skin_id(ws, col_id, hero_id):
    """Find the maximum skin ID for a given hero in Item.xlsx HeroSkin rows."""
    max_id = 0
    hero_prefix = str(hero_id)
    for row in range(5, ws.max_row + 1):
        c3 = ws.cell(row=row, column=3).value  # type column
        c1 = ws.cell(row=row, column=1).value  # id column
        if c3 == "HeroSkin" and c1 is not None:
            c1_str = str(c1)
            if c1_str.startswith(hero_prefix):
                try:
                    max_id = max(max_id, int(c1_str))
                except ValueError:
                    pass
    return max_id


def find_hero_data(ws, hero_name):
    """Find hero by name in Hero.xlsx. Returns (hero_id, hero_pinyin) or (None, None)."""
    for row in range(5, ws.max_row + 1):
        name = ws.cell(row=row, column=2).value
        if name is not None and str(name).strip() == hero_name:
            hero_id = ws.cell(row=row, column=1).value
            hero_pinyin = ws.cell(row=row, column=3).value
            return hero_id, hero_pinyin
    return None, None


def find_skin_lines(ws, skin_name):
    """
    Find all rows in HeroLines where col2(所属皮肤) == skin_name.
    Returns dict with: all_lines, debut_lines, kill_lines, dead_lines, opt_lines
    Each value is a comma-separated string of line IDs.
    """
    debut_ids = []
    kill_ids = []
    dead_ids = []
    opt_ids = []
    all_ids = []

    for row in range(5, ws.max_row + 1):
        skin = ws.cell(row=row, column=2).value
        if skin is None or str(skin).strip() != skin_name:
            continue

        line_id = ws.cell(row=row, column=1).value
        if line_id is None:
            continue
        line_id_str = str(line_id)
        line_id_int = int(line_id_str)

        line_type = ws.cell(row=row, column=3).value
        tag = ws.cell(row=row, column=4).value if ws.max_column >= 4 else None

        line_type_str = str(line_type).strip() if line_type else ""
        tag_str = str(tag).strip() if tag else ""

        if line_type_str == "LinesType_Dengchang" or tag_str == "登场":
            debut_ids.append(line_id_int)
        elif line_type_str == "LinesType_Kill" or tag_str == "击杀":
            kill_ids.append(line_id_int)
        elif line_type_str == "LinesType_Dead" or tag_str == "阵亡":
            dead_ids.append(line_id_int)
        elif tag_str == "自选":
            opt_ids.append(line_id_int)
        else:
            all_ids.append(line_id_int)

    # Sort numerically before joining (each group and the combined all_lines)
    all_ids.sort()
    debut_ids.sort()
    kill_ids.sort()
    dead_ids.sort()
    opt_ids.sort()

    # all_lines = global sorted order of all non-自选 IDs
    combined_non_opt = sorted(all_ids + debut_ids + kill_ids + dead_ids)

    return {
        "all_lines": ",".join(str(x) for x in combined_non_opt) if combined_non_opt else "",
        "debut_lines": ",".join(str(x) for x in debut_ids),
        "kill_lines": ",".join(str(x) for x in kill_ids),
        "dead_lines": ",".join(str(x) for x in dead_ids),
        "opt_lines": ",".join(str(x) for x in opt_ids),
    }


# ──────────────────────────────────────────────────────────────────────
# Step implementations
# ──────────────────────────────────────────────────────────────────────

def step1_hero(base_dir, hero_name):
    """Step 1: Hero.xlsx - Find hero by name, get HeroID and HeroPinyin."""
    hero_path = os.path.join(base_dir, "Hero.xlsx")
    if not os.path.exists(hero_path):
        print(f"ERROR: Hero.xlsx not found at {hero_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(hero_path)
    ws = wb["武将表|Hero"]

    hero_id, hero_pinyin = find_hero_data(ws, hero_name)
    wb.close()

    if hero_id is None:
        print(f"ERROR: Hero '{hero_name}' not found in Hero.xlsx")
        sys.exit(1)

    hero_item_id = int("10" + str(hero_id))
    hero_pinyin_lower = str(hero_pinyin).lower() if hero_pinyin else ""

    print(f"[Step 1 - Hero.xlsx]")
    print(f"  HeroName: {hero_name}")
    print(f"  HeroID: {hero_id}")
    print(f"  HeroPinyin: {hero_pinyin}")
    print(f"  HeroPinyin (lower): {hero_pinyin_lower}")
    print(f"  HeroItemId: {hero_item_id}")

    return hero_id, hero_pinyin, hero_pinyin_lower, hero_item_id


def step2_item(base_dir, hero_name, skin_name, skin_pinyin, level, hero_id,
               hero_pinyin_lower, hero_item_id):
    """Step 2: Item.xlsx - Clone and modify a new HeroSkin row."""
    item_path = os.path.join(base_dir, "Item.xlsx")
    if not os.path.exists(item_path):
        print(f"ERROR: Item.xlsx not found at {item_path}")
        sys.exit(1)

    quality = LEVEL_MAP_QUALITY[level]

    wb = openpyxl.load_workbook(item_path)
    ws = wb["道具表|Item"]

    # Find max skin ID for this hero
    max_skin_id = find_max_skin_id(ws, 1, hero_id)
    if max_skin_id == 0:
        print(f"ERROR: No existing HeroSkin rows found for hero_id={hero_id} in Item.xlsx")
        wb.close()
        sys.exit(1)

    new_skin_id = max_skin_id + 1
    print(f"\n[Step 2 - Item.xlsx]")
    print(f"  Max SkinID: {max_skin_id}")
    print(f"  New SkinID: {new_skin_id}")

    # Find the row with max_skin_id to use as reference for cloning
    ref_row = None
    for row in range(5, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1).value
        if c1 is not None and str(c1) == str(max_skin_id):
            ref_row = row
            break

    if ref_row is None:
        print(f"ERROR: Cannot find row with SkinID={max_skin_id} to clone")
        wb.close()
        sys.exit(1)

    max_col = ws.max_column
    new_row = insert_row_below(ws, ref_row, max_col)
    print(f"  Cloned row {ref_row} → new row {new_row}")

    # Modify the new row
    ws.cell(row=new_row, column=1).value = new_skin_id
    ws.cell(row=new_row, column=2).value = skin_name
    ws.cell(row=new_row, column=4).value = quality
    ws.cell(row=new_row, column=12).value = hero_item_id

    icon_path = f"UI/Images/Item/ui_s1_daoju_wujiang_{hero_pinyin_lower}_{skin_pinyin}.png"
    ws.cell(row=new_row, column=22).value = icon_path

    desc = f"{hero_name}的武将皮肤，获得后可以在游戏中改变武将{hero_name}的卡牌形象"
    ws.cell(row=new_row, column=28).value = desc

    print(f"  col1  (ID):        {new_skin_id}")
    print(f"  col2  (Name):      {skin_name}")
    print(f"  col4  (Quality):   {quality}")
    print(f"  col12 (HeroItemId):{hero_item_id}")
    print(f"  col22 (Icon):      {icon_path}")
    print(f"  col28 (Desc):      {desc}")

    wb.save(item_path)
    wb.close()
    print(f"  ✅ Item.xlsx saved")

    return new_skin_id


def step3_heroskinitem(base_dir, skin_name, level, hero_id, hero_pinyin_lower,
                       skin_pinyin, new_skin_id):
    """Step 3: HeroSkinItem_英雄皮肤.xlsx - Clone and modify a new skin row."""
    hs_path = os.path.join(base_dir, "HeroSkinItem_英雄皮肤.xlsx")
    if not os.path.exists(hs_path):
        print(f"ERROR: HeroSkinItem_英雄皮肤.xlsx not found at {hs_path}")
        sys.exit(1)

    raily_type = LEVEL_MAP_RAILY[level]

    wb = openpyxl.load_workbook(hs_path)
    ws = wb["英雄皮肤|HeroSkinItem"]
    max_col = ws.max_column

    # Find the last row with a skin for this hero (col2 == hero_id)
    ref_row = None
    for row in range(5, ws.max_row + 1):
        c2 = ws.cell(row=row, column=2).value
        if c2 is not None and str(c2) == str(hero_id):
            ref_row = row
    # Fallback: use the last row with any data
    if ref_row is None:
        for row in range(ws.max_row, 4, -1):
            c1 = ws.cell(row=row, column=1).value
            if c1 is not None:
                ref_row = row
                break
    if ref_row is None:
        ref_row = 4

    new_row = insert_row_below(ws, ref_row, max_col)
    print(f"\n[Step 3 - HeroSkinItem_英雄皮肤.xlsx]")
    print(f"  Cloned row {ref_row} → new row {new_row}")

    skin_pinyin_full = f"{hero_pinyin_lower}_{skin_pinyin}"

    ws.cell(row=new_row, column=1).value = new_skin_id
    ws.cell(row=new_row, column=2).value = hero_id
    ws.cell(row=new_row, column=3).value = skin_name
    ws.cell(row=new_row, column=4).value = raily_type
    ws.cell(row=new_row, column=5).value = "SkinOtherSkin"
    ws.cell(row=new_row, column=6).value = skin_name
    ws.cell(row=new_row, column=7).value = "待配置"
    ws.cell(row=new_row, column=8).value = skin_pinyin_full
    ws.cell(row=new_row, column=9).value = None
    ws.cell(row=new_row, column=10).value = None
    ws.cell(row=new_row, column=11).value = None  # fill in Step 4
    ws.cell(row=new_row, column=12).value = None
    ws.cell(row=new_row, column=13).value = None
    ws.cell(row=new_row, column=14).value = None
    ws.cell(row=new_row, column=15).value = None
    ws.cell(row=new_row, column=16).value = "待配置"
    ws.cell(row=new_row, column=17).value = "待配置"
    ws.cell(row=new_row, column=18).value = "待配置"
    ws.cell(row=new_row, column=19).value = 0

    print(f"  SkinID={new_skin_id} | HeroID={hero_id} | Raily={raily_type}")
    print(f"  Name={skin_name} | Pinyin={skin_pinyin_full} | SkinType=SkinOtherSkin")

    wb.save(hs_path)
    wb.close()
    print(f"  ✅ HeroSkinItem_英雄皮肤.xlsx saved (台词等待 Step 4)")


def step4_hero_lines(base_dir, skin_name, new_skin_id):
    """Step 4: HeroLines_武将台词表.xlsx - Find lines and fill HeroSkinItem."""
    hl_path = os.path.join(base_dir, "HeroLines_武将台词表.xlsx")
    hs_path = os.path.join(base_dir, "HeroSkinItem_英雄皮肤.xlsx")

    if not os.path.exists(hs_path):
        print(f"ERROR: HeroSkinItem_英雄皮肤.xlsx not found for update")
        sys.exit(1)

    print(f"\n[Step 4 - HeroLines_武将台词表.xlsx]")

    if not os.path.exists(hl_path):
        print(f"  WARNING: HeroLines_武将台词表.xlsx not found, lines left unconfigured")
        return

    # Read lines
    wb = openpyxl.load_workbook(hl_path)
    ws = wb["武将台词|HeroLines"]
    line_data = find_skin_lines(ws, skin_name)
    wb.close()

    print(f"  all_lines   (全部非自选): {line_data['all_lines']}")
    print(f"  debut_lines (登场):       {line_data['debut_lines']}")
    print(f"  kill_lines  (击杀):       {line_data['kill_lines']}")
    print(f"  dead_lines  (阵亡):       {line_data['dead_lines']}")
    print(f"  opt_lines   (自选):       {line_data['opt_lines']}")

    # Update HeroSkinItem voice line columns
    wb = openpyxl.load_workbook(hs_path)
    ws = wb["英雄皮肤|HeroSkinItem"]

    # Find the row with our new SkinID
    target_row = None
    for row in range(5, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1).value
        if c1 is not None and str(c1) == str(new_skin_id):
            target_row = row
            break

    if target_row is None:
        print(f"  ERROR: Cannot find row with SkinID={new_skin_id} in HeroSkinItem")
        wb.close()
        sys.exit(1)

    # Set number_format='@' (text) BEFORE setting values to preserve as strings
    voice_cols = {
        11: ("武将台词", line_data["all_lines"]),
        12: ("武将登场台词", line_data["debut_lines"]),
        13: ("武将击杀台词", line_data["kill_lines"]),
        14: ("武将阵亡台词", line_data["dead_lines"]),
        15: ("武将自选台词", line_data["opt_lines"]),
    }

    for col, (label, value) in voice_cols.items():
        cell = ws.cell(row=target_row, column=col)
        cell.number_format = '@'  # text format
        cell.value = value if value else ""
        print(f"  col{col} ({label}): {value if value else '(空)'}")

    wb.save(hs_path)
    wb.close()
    print(f"  ✅ HeroSkinItem_英雄皮肤.xlsx updated with voice lines")


def step5_heroskinspine(base_dir, skin_name, new_skin_id, hero_pinyin, hero_pinyin_lower):
    """Step 5: HeroSkinSpine_英雄皮肤Spine.xlsx - Clone and modify spine config.

    Insert new row above "#动态皮肤", copy template row above insertion point,
    then replace hero name in SpineAnimAudio / KillAudio / SpineAnimBattleAudio.
    """
    import re as _re

    sp_path = os.path.join(base_dir, "HeroSkinSpine_英雄皮肤Spine.xlsx")
    if not os.path.exists(sp_path):
        print(f"WARNING: HeroSkinSpine_英雄皮肤Spine.xlsx not found, skipping")
        return

    hero_pascal = str(hero_pinyin) if hero_pinyin else ""  # PascalCase from Hero.xlsx col3

    wb = openpyxl.load_workbook(sp_path)
    ws = wb["英雄皮肤Spine|HeroSkinSpine"]
    max_col = ws.max_column

    # Find "#动态皮肤" marker row
    marker_row = None
    for row in range(5, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1).value
        if c1 is not None and str(c1).strip() == "#动态皮肤":
            marker_row = row
            break

    if marker_row is None:
        print(f"  ERROR: Cannot find '#动态皮肤' marker row in HeroSkinSpine")
        wb.close()
        sys.exit(1)

    # Template row is the one immediately above the marker
    template_row = marker_row - 1

    # Insert new row at marker position (pushes marker down)
    safe_insert_rows(ws, marker_row, 1)
    new_row = marker_row

    # Clone template row (copy values + styles)
    for col in range(1, max_col + 1):
        src = ws.cell(row=template_row, column=col)
        dst = ws.cell(row=new_row, column=col)
        dst.value = src.value
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = copy(src.number_format)
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)

    print(f"\n[Step 5 - HeroSkinSpine_英雄皮肤Spine.xlsx]")
    print(f"  Inserted above '#动态皮肤' (was row {marker_row})")
    print(f"  Cloned template row {template_row} → new row {new_row}")

    # Read template audio values
    template_c11 = ws.cell(row=template_row, column=11).value or ""
    template_c13 = ws.cell(row=template_row, column=13).value or ""
    template_c14 = ws.cell(row=template_row, column=14).value or ""
    print(f"  Template audio: col11 HeroPascal in pattern")

    # Modify new row
    ws.cell(row=new_row, column=1).value = new_skin_id
    ws.cell(row=new_row, column=2).value = skin_name

    # Replace hero name (PascalCase) in audio config strings
    # Pattern: Character_{HeroPascal}_ or Character_{HeroPascal}Skin_
    def replace_hero_in_audio(audio_str):
        if not audio_str:
            return audio_str
        # Match Character_ followed by PascalCase name (may have "Skin" suffix)
        # PascalCase: one or more [A-Z][a-z]+ segments, e.g. "RuJi", "SunQuan"
        return _re.sub(
            r'Character_([A-Z][a-z]+(?:[A-Z][a-z]+)*)(Skin)?_',
            rf'Character_{hero_pascal}\2_',
            str(audio_str)
        )

    ws.cell(row=new_row, column=11).value = replace_hero_in_audio(template_c11)
    ws.cell(row=new_row, column=13).value = replace_hero_in_audio(template_c13)
    ws.cell(row=new_row, column=14).value = replace_hero_in_audio(template_c14)

    print(f"  col1  (SkinID):     {new_skin_id}")
    print(f"  col2  (Name):       {skin_name}")
    print(f"  col11 (Audio):      {ws.cell(row=new_row, column=11).value}")
    print(f"  col13 (KillAudio):  {ws.cell(row=new_row, column=13).value}")
    print(f"  col14 (BattleAudio):{ws.cell(row=new_row, column=14).value}")

    wb.save(sp_path)
    wb.close()
    print(f"  ✅ HeroSkinSpine_英雄皮肤Spine.xlsx saved")


def step6_itemheroskin(base_dir, skin_name, new_skin_id, hero_pinyin_lower):
    """Step 6: ItemHeroSkin_武将皮肤展示表.xlsx - Clone and modify display config.

    Configures path columns by replacing hero pinyin in the pattern:
      ui_s1_gongxihuode_bg_{hero_lower} / ui_s1_fight_victory_{hero_lower}_bg / ...
    坐标和音效列不配置（留空）。
    """
    import re as _re

    ihs_path = os.path.join(base_dir, "ItemHeroSkin_武将皮肤展示表.xlsx")
    if not os.path.exists(ihs_path):
        print(f"WARNING: ItemHeroSkin_武将皮肤展示表.xlsx not found, skipping")
        return

    wb = openpyxl.load_workbook(ihs_path)
    ws = wb["武将皮肤展示表|ItemHeroSkin"]
    max_col = ws.max_column

    # Find last row with data
    ref_row = None
    for row in range(ws.max_row, 4, -1):
        c1 = ws.cell(row=row, column=1).value
        if c1 is not None:
            ref_row = row
            break
    if ref_row is None:
        ref_row = 4

    # Read template path values
    template_col3 = ws.cell(row=ref_row, column=3).value or ""
    template_col4 = ws.cell(row=ref_row, column=4).value or ""
    template_col5 = ws.cell(row=ref_row, column=5).value or ""
    template_col7 = ws.cell(row=ref_row, column=7).value or ""

    # Detect old hero pinyin from path patterns
    # Pattern: ui_s1_gongxihuode_bg_{hero_lower}  or  ui_s1_fight_victory_{hero_lower}_bg
    old_hp = ""
    for path_str in [template_col3, template_col4, template_col5, template_col7]:
        if not path_str:
            continue
        m = _re.search(r'(?:bg_|victory_)([a-z]+)(?:_bg|_role|_text)?$', str(path_str))
        if m:
            old_hp = m.group(1)
            break

    new_row = insert_row_below(ws, ref_row, max_col)
    print(f"\n[Step 6 - ItemHeroSkin_武将皮肤展示表.xlsx]")
    print(f"  Cloned row {ref_row} → new row {new_row}")
    print(f"  Detected old hero pinyin: '{old_hp}' → new: '{hero_pinyin_lower}'")

    # Replace hero pinyin in path strings
    def replace_pinyin(s):
        if not s or not old_hp:
            return s
        return str(s).replace(old_hp, hero_pinyin_lower)

    ws.cell(row=new_row, column=1).value = new_skin_id
    ws.cell(row=new_row, column=2).value = skin_name
    ws.cell(row=new_row, column=3).value = replace_pinyin(template_col3)
    ws.cell(row=new_row, column=4).value = replace_pinyin(template_col4)
    ws.cell(row=new_row, column=5).value = replace_pinyin(template_col5)
    ws.cell(row=new_row, column=6).value = None   # 立绘坐标 不配置
    ws.cell(row=new_row, column=7).value = replace_pinyin(template_col7)
    ws.cell(row=new_row, column=8).value = None   # 文本坐标 不配置
    ws.cell(row=new_row, column=9).value = None   # 音效 不配置

    print(f"  col1 (SkinID):     {new_skin_id}")
    print(f"  col2 (Name):       {skin_name}")
    print(f"  col3 (立绘路径):    {ws.cell(row=new_row, column=3).value}")
    print(f"  col4 (背景):        {ws.cell(row=new_row, column=4).value}")
    print(f"  col5 (立绘):        {ws.cell(row=new_row, column=5).value}")
    print(f"  col6 (坐标):        不配置")
    print(f"  col7 (文本图):      {ws.cell(row=new_row, column=7).value}")
    print(f"  col8 (文本坐标):    不配置")
    print(f"  col9 (音效):        不配置")

    wb.save(ihs_path)
    wb.close()
    print(f"  ✅ ItemHeroSkin_武将皮肤展示表.xlsx saved")


# ──────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Configure a new hero skin item across multiple Excel files."
    )
    parser.add_argument("base_dir", help="Base config directory (e.g. /path/to/名将杀配置/)")
    parser.add_argument("skin_name", help="皮肤名称 (e.g. 驰猎惊澜)")
    parser.add_argument("hero_name", help="武将名 (e.g. 孙权)")
    parser.add_argument("skin_pinyin", help="皮肤拼音 (e.g. chiliejinglan)")
    parser.add_argument("level", choices=["精良", "卓越", "至臻"], help="皮肤品质等级")

    args = parser.parse_args()

    if args.level not in LEVEL_MAP_QUALITY:
        print(f"ERROR: Invalid level '{args.level}', must be one of: 精良, 卓越, 至臻")
        sys.exit(1)

    print("=" * 60)
    print(f" 配置英雄皮肤: {args.skin_name} ({args.hero_name})")
    print(f" 等级: {args.level} | 拼音: {args.skin_pinyin}")
    print(f" 配置目录: {args.base_dir}")
    print("=" * 60)

    # Step 1: Hero.xlsx
    hero_id, hero_pinyin, hero_pinyin_lower, hero_item_id = step1_hero(
        args.base_dir, args.hero_name
    )

    # Step 2: Item.xlsx
    new_skin_id = step2_item(
        args.base_dir, args.hero_name, args.skin_name, args.skin_pinyin,
        args.level, hero_id, hero_pinyin_lower, hero_item_id
    )

    # Step 3: HeroSkinItem_英雄皮肤.xlsx
    step3_heroskinitem(
        args.base_dir, args.skin_name, args.level, hero_id,
        hero_pinyin_lower, args.skin_pinyin, new_skin_id
    )

    # Step 4: HeroLines + fill HeroSkinItem voice line columns
    step4_hero_lines(args.base_dir, args.skin_name, new_skin_id)

    # Step 5: HeroSkinSpine_英雄皮肤Spine.xlsx
    step5_heroskinspine(
        args.base_dir, args.skin_name, new_skin_id, hero_pinyin, hero_pinyin_lower
    )

    # Step 6: ItemHeroSkin_武将皮肤展示表.xlsx
    step6_itemheroskin(args.base_dir, args.skin_name, new_skin_id, hero_pinyin_lower)

    print("")
    print("=" * 60)
    print(f" ✅ 配置完成!")
    print(f"    新皮肤 ID: {new_skin_id}")
    print(f"    皮肤名称:  {args.skin_name}")
    print(f"    品质等级:  {args.level}")
    print("=" * 60)

    return new_skin_id


if __name__ == "__main__":
    main()
