#!/usr/bin/env python3
"""
名将杀 - 伙伴道具配置脚本
用法: python3 configure_hero_companion.py <伙伴名> [选项]
示例: python3 configure_hero_companion.py 麒麟

流程:
1. 从飞书 Wiki 读取伙伴信息（品质/属性/技能名称/时间）
2. 从 Skill.xlsx #伙伴技能 区域查找技能ID
3. Item.xlsx 道具表|Item: 新增 Partner 类型道具
4. Item.xlsx 形象表|ImageItem: 新增伙伴形象
5. Pet_灵宠表.xlsx: 新增灵宠条目
"""

import sys
import subprocess
import json
import os
from datetime import datetime, timedelta
from copy import copy

# === 配置常量 ===
PYTHON_VENV = os.path.expanduser(
    "~/.openclaw/skills/aicconfig/scripts/aicconfig/venv/bin/python3"
)
EXCEL_DIR = os.path.expanduser("~/.openclaw/workspace/名将杀配置")

# Wiki 数据源
# 伙伴规划 (品质/拼音/投放时间)
PLAN_SPREADSHEET = "PdtjsYqxnh7iybtISNecV2dZneb"
PLAN_SHEET = "4pfFqd"  # 伙伴规划
# 伙伴养成和技能数值 (属性/技能名称)
SKILL_WIKI_SPREADSHEET = "YjDestOAchLLFSt80kicLcnynmb"
SKILL_WIKI_SHEET = "q6Dunq"  # 每个伙伴维度

# 品级 → 道具品质(Rarity) 映射
QUALITY_MAP = {
    "传说": 4,
    "史诗": 3,
    "稀有": 2,
    "普通": 1,
}


def lark_cli(*args):
    """调用 lark-cli 并返回 JSON"""
    cmd = ["lark-cli"] + list(args)
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    if result.returncode != 0:
        print(f"lark-cli 错误: {result.stderr}")
        return None
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError:
        print(f"lark-cli 返回非 JSON: {result.stdout[:200]}")
        return None


def get_companion_info_from_wiki(companion_name):
    """
    Step 1: 从飞书 Wiki 读取伙伴信息
    - 伙伴规划 → 品质/拼音/投放时间
    - 伙伴养成和技能数值 → 属性/技能名称
    """
    print(f"🔍 在飞书 Wiki 中搜索伙伴: {companion_name}")

    # --- 读取伙伴规划 (品质/拼音/投放时间) ---
    result = lark_cli(
        "sheets", "+find",
        "--spreadsheet-token", PLAN_SPREADSHEET,
        "--sheet-id", PLAN_SHEET,
        "--find", companion_name,
    )
    if not result or not result.get("ok"):
        print("  ❌ 伙伴规划查询失败")
        return None
    cells = result.get("data", {}).get("find_result", {}).get("matched_cells", [])
    if not cells:
        print(f"  ❌ 伙伴规划中未找到: {companion_name}")
        return None

    plan_row = int(cells[0][1:])
    result = lark_cli(
        "sheets", "+read",
        "--spreadsheet-token", PLAN_SPREADSHEET,
        "--sheet-id", PLAN_SHEET,
        "--range", f"{PLAN_SHEET}!B{plan_row}:F{plan_row}",
    )
    if not result or not result.get("ok"):
        print("  ❌ 读取伙伴规划行失败")
        return None
    plan_vals = result["data"]["valueRange"]["values"][0]
    plan_name = plan_vals[0] if len(plan_vals) > 0 else None
    plan_pinyin = plan_vals[1] if len(plan_vals) > 1 else None
    plan_quality = plan_vals[2] if len(plan_vals) > 2 else None
    plan_time_num = plan_vals[3] if len(plan_vals) > 3 else None  # Excel serial
    plan_method = plan_vals[4] if len(plan_vals) > 4 else None

    # 转换时间
    start_time = None
    if plan_time_num:
        try:
            base = datetime(1899, 12, 30)
            start_time = (base + timedelta(days=int(plan_time_num))).strftime("%Y-%m-%d 00:00:00")
        except:
            pass

    pinyin = plan_pinyin or companion_name.lower()
    print(f"  规划: name={plan_name}, pinyin={pinyin}, quality={plan_quality}, time={start_time}")

    # --- 读取伙伴养成和技能数值 (属性/技能名称) ---
    result = lark_cli(
        "sheets", "+find",
        "--spreadsheet-token", SKILL_WIKI_SPREADSHEET,
        "--sheet-id", SKILL_WIKI_SHEET,
        "--find", companion_name,
    )
    skill1_name = None
    skill2_name = None
    if result and result.get("ok"):
        cells = result.get("data", {}).get("find_result", {}).get("matched_cells", [])
        if cells:
            skill_row = int(cells[0][1:])
            result = lark_cli(
                "sheets", "+read",
                "--spreadsheet-token", SKILL_WIKI_SPREADSHEET,
                "--sheet-id", SKILL_WIKI_SHEET,
                "--range", f"{SKILL_WIKI_SHEET}!A{skill_row}:O{skill_row}",
            )
            if result and result.get("ok"):
                row = result["data"]["valueRange"]["values"][0]
                skill1_name = row[11] if len(row) > 11 and row[11] else None  # L
                skill2_name = row[13] if len(row) > 13 and row[13] else None  # N

    info = {
        "name": companion_name,
        "pinyin": pinyin,
        "quality": plan_quality or "稀有",
        "start_time": start_time,
        "skill_names": [n for n in [skill1_name, skill2_name] if n],
        "prefab_path": f"Prefabs/Pet/Pet_{companion_name}.prefab",
        "icon_path": f"UI/Images/Pet/ui_s1_pet_head_{pinyin}_01.png",
        "square_icon": f"UI/Images/Pet/Square/ui_s1_touxiang_pet_{pinyin}.png",
        "silhouette": f"UI/Images/Pet/ui_s1_pet_role_{pinyin}.png",
    }
    print(f"  📋 品质={info['quality']}, 时间={info['start_time']}, 技能={info['skill_names']}")
    return info


def find_skill_ids(skill_names):
    """Step 2: 从 Skill.xlsx #伙伴技能 区域查找技能ID"""
    import openpyxl
    print(f"\n🔍 在 Skill.xlsx 中查找技能ID: {skill_names}")
    skill_path = os.path.join(EXCEL_DIR, "Skill.xlsx")
    wb = openpyxl.load_workbook(skill_path, data_only=True)
    ws = wb["技能表|Skill"]
    skill_map = {}
    for row_idx in range(751, ws.max_row + 1):
        sn = str(ws.cell(row_idx, 2).value or "")
        si = ws.cell(row_idx, 1).value
        if sn and si:
            skill_map[sn] = int(si)
    wb.close()
    result = []
    for name in skill_names:
        sid = skill_map.get(name)
        if sid:
            print(f"  ✅ {name} → {sid}")
            result.append({"name": name, "skill_id": sid})
        else:
            print(f"  ⚠️ 未找到: {name}")
    return result


def configure_item(companion_info):
    """Step 3: 在 Item.xlsx 道具表|Item 中新增 Partner 道具"""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    item_path = os.path.join(EXCEL_DIR, "Item.xlsx")
    print(f"\n📂 打开 Item.xlsx")
    wb = openpyxl.load_workbook(item_path)
    ws = wb["道具表|Item"]

    # 找到 Partner 类型道具区域的最大 ID
    max_id = 0
    last_partner_row = 0
    for row_idx in range(5, ws.max_row + 1):
        type_val = ws.cell(row_idx, 3).value  # C - Type
        id_val = ws.cell(row_idx, 1).value    # A - Id
        if type_val and str(type_val).strip() == "Partner":
            if id_val and str(id_val).strip().isdigit():
                max_id = max(max_id, int(id_val))
                last_partner_row = row_idx

    new_id = max_id + 1
    print(f"  当前最大 Partner ID: {max_id}, 新 ID: {new_id}")

    # 在最后一行下方插入
    insert_row = last_partner_row + 1
    ws.insert_rows(insert_row)
    print(f"  在 Row {insert_row} 插入新行")

    # 复制上一行样式和值
    src_row = insert_row - 1
    max_col = 37  # Item 表有效列数

    for col in range(1, max_col + 1):
        src_cell = ws.cell(src_row, col)
        dst_cell = ws.cell(insert_row, col)

        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)
            dst_cell.alignment = copy(src_cell.alignment)

        # 复制值（除了需要覆盖的列）
        if col not in [1, 2, 3, 4, 22, 24, 25, 28, 29, 36, 37]:
            dst_cell.value = src_cell.value

    name = companion_info["name"]
    pinyin = companion_info.get("pinyin", name.lower())
    quality_num = QUALITY_MAP.get(companion_info.get("quality", "稀有"), 2)

    # 写入新值
    ws.cell(insert_row, 1).value = new_id               # Id
    ws.cell(insert_row, 2).value = name                  # Name
    ws.cell(insert_row, 3).value = "Partner"             # Type
    ws.cell(insert_row, 4).value = quality_num           # Rarity
    ws.cell(insert_row, 22).value = companion_info.get("icon_path", "")  # Icon
    ws.cell(insert_row, 24).value = 1                    # IsGetHint
    ws.cell(insert_row, 25).value = 1                    # IsHide
    ws.cell(insert_row, 28).value = "待配置"              # Des
    ws.cell(insert_row, 29).value = companion_info.get("description", "待配置")  # DesB
    ws.cell(insert_row, 36).value = companion_info.get("display_icon", "")  # DisplayIcon
    ws.cell(insert_row, 37).value = 1                    # HaveCheck

    print(f"  ✅ Item Row {insert_row}: Id={new_id}, Name={name}, Type=Partner, Rarity={quality_num}")
    wb.save(item_path)
    wb.close()

    return new_id


def configure_image_item(companion_info, item_id):
    """Step 4: 在 Item.xlsx 形象表|ImageItem 中新增形象"""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    item_path = os.path.join(EXCEL_DIR, "Item.xlsx")
    print(f"\n📂 打开 Item.xlsx (形象表|ImageItem)")
    wb = openpyxl.load_workbook(item_path)
    ws = wb["形象表|ImageItem"]

    # 找到最大形象 ID (50000xx 范围)
    max_img_id = 0
    last_img_row = 0
    for row_idx in range(5, ws.max_row + 1):
        id_val = ws.cell(row_idx, 1).value
        if id_val and str(id_val).strip().isdigit():
            id_num = int(id_val)
            if 5000000 <= id_num <= 5999999:
                if id_num > max_img_id:
                    max_img_id = id_num
                    last_img_row = row_idx

    # 形象 ID 与道具 ID 相同
    img_id = item_id
    print(f"  形象 ID: {img_id}, 参考行: Row {last_img_row}")

    # 插入
    insert_row = last_img_row + 1
    ws.insert_rows(insert_row)
    print(f"  在 Row {insert_row} 插入新行")

    # 复制样式
    src_row = insert_row - 1
    for col in range(1, 12):  # ImageItem 有 11 列
        src_cell = ws.cell(src_row, col)
        dst_cell = ws.cell(insert_row, col)

        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)
            dst_cell.alignment = copy(src_cell.alignment)

        if col not in [1, 2, 5]:
            dst_cell.value = src_cell.value

    name = companion_info["name"]
    pinyin = companion_info.get("pinyin", name.lower())

    ws.cell(insert_row, 1).value = img_id                   # Id
    ws.cell(insert_row, 2).value = name                      # Name
    ws.cell(insert_row, 5).value = companion_info.get(
        "head_icon",
        f"UI/Images/Global/touxiang/ui_s1_touxiang_zhujiemian_{pinyin}.png"
    )

    print(f"  ✅ ImageItem Row {insert_row}: Id={img_id}, Name={name}")
    wb.save(item_path)
    wb.close()


def configure_pet(companion_info, item_id):
    """Step 5: 在 Pet_灵宠表.xlsx 中新增灵宠条目"""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    pet_path = os.path.join(EXCEL_DIR, "Pet_灵宠表.xlsx")
    print(f"\n📂 打开 Pet_灵宠表.xlsx")
    wb = openpyxl.load_workbook(pet_path)
    ws = wb["灵宠|Pet"]

    # 找到最大灵宠 ID
    max_pet_id = 0
    last_pet_row = 0
    for row_idx in range(5, ws.max_row + 1):
        id_val = ws.cell(row_idx, 1).value
        if id_val and str(id_val).strip().isdigit():
            id_num = int(id_val)
            if id_num > max_pet_id:
                max_pet_id = id_num
                last_pet_row = row_idx

    pet_id = item_id  # 与道具 ID 一致
    print(f"  当前最大 Pet ID: {max_pet_id}, 新 Pet ID: {pet_id}")

    # 插入
    insert_row = last_pet_row + 1
    ws.insert_rows(insert_row)
    print(f"  在 Row {insert_row} 插入新行")

    # 复制样式
    src_row = insert_row - 1
    for col in range(1, 23):  # Pet 表有 22 列
        src_cell = ws.cell(src_row, col)
        dst_cell = ws.cell(insert_row, col)

        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)
            dst_cell.alignment = copy(src_cell.alignment)

        # 只覆盖差异列，其余从上一行继承
        if col not in [1, 2, 4, 5, 6, 7, 11, 12, 13, 22]:
            dst_cell.value = src_cell.value

    name = companion_info["name"]
    pinyin = companion_info.get("pinyin", name.lower())

    # 技能格式: {技能槽位;技能ID}
    skills = companion_info.get("skills", [])
    skill_str = ",".join([f"{{{s['slot']};{s['skill_id']}}}" for s in skills]) if skills else ""

    ws.cell(insert_row, 1).value = pet_id                    # Id
    ws.cell(insert_row, 2).value = name                       # Name
    ws.cell(insert_row, 4).value = companion_info.get("prefab_path", "")  # PrefabPath
    ws.cell(insert_row, 5).value = companion_info.get("start_time", "")   # StartTime
    ws.cell(insert_row, 6).value = skill_str                  # Skill
    ws.cell(insert_row, 7).value = companion_info.get("battle_attr_weight", "")  # BattleAttrWeight
    ws.cell(insert_row, 11).value = companion_info.get("square_icon", "")   # SquareHeadIcon
    ws.cell(insert_row, 12).value = companion_info.get("icon_path", "")      # HeadIcon
    ws.cell(insert_row, 13).value = companion_info.get("silhouette", "")     # Silhouette
    ws.cell(insert_row, 22).value = "待配置"                   # InfoTextID

    print(f"  ✅ Pet Row {insert_row}: Id={pet_id}, Name={name}, Skills={skill_str}")
    wb.save(pet_path)
    wb.close()


def main():
    if len(sys.argv) < 2:
        print("用法: python3 configure_hero_companion.py <伙伴名>")
        print("示例: python3 configure_hero_companion.py 麒麟")
        sys.exit(1)

    companion_name = sys.argv[1]
    print(f"🐙 开始配置伙伴道具: {companion_name}")
    print("=" * 50)

    # Step 1: 从 Wiki 获取伙伴信息
    companion_info = get_companion_info_from_wiki(companion_name)
    if not companion_info:
        sys.exit(1)

    # Step 2: 查找技能ID
    skill_ids = find_skill_ids(companion_info.get("skill_names", []))
    companion_info["skills"] = skill_ids

    # Step 3: 配置 Item.xlsx 道具表|Item
    item_id = configure_item(companion_info)

    # Step 4: 配置 Item.xlsx 形象表|ImageItem
    configure_image_item(companion_info, item_id)

    # Step 5: 配置 Pet_灵宠表.xlsx
    configure_pet(companion_info, item_id)

    print(f"\n🎉 完成! 伙伴 {companion_name} 道具已配置完成")
    print(f"   道具ID: {item_id}")
    print(f"   涉及文件: Item.xlsx (道具表+形象表), Pet_灵宠表.xlsx")


if __name__ == "__main__":
    main()
