#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
名将杀 skill-tester：将 report_data.json 导出为 Excel。

推荐通过统一入口调用:
  python skill_test_report.py export --skill-name "宫" --input report_data.json

也可直接:
  python export_test_report.py --skill-dir "../宫" --input report_data.json

依赖: openpyxl (pip install openpyxl)
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要安装 openpyxl，请执行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

from skill_test_common import merge_report_meta, normalize_result, resolve_testcase_md, row_from_obj

HEADER = ["用例ID", "标题", "前置条件", "预期", "实际", "结果", "规则依据（代码路径）"]
COL_WIDTHS = [10, 28, 36, 36, 36, 10, 42]

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PASS_FILL = PatternFill("solid", fgColor="E2EFDA")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PENDING_FILL = PatternFill("solid", fgColor="FFF2CC")


def style_header_row(ws, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def write_detail_sheet(ws, rows: list[tuple]) -> None:
    ws.title = "测试明细"
    for i, h in enumerate(HEADER, 1):
        ws.cell(1, i, h)
    style_header_row(ws, len(HEADER))

    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.alignment = WRAP
            cell.border = BORDER
            if c == 6:
                if val == "通过":
                    cell.fill = PASS_FILL
                elif val == "失败":
                    cell.fill = FAIL_FILL
                elif val == "待验证":
                    cell.fill = PENDING_FILL

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADER))}{len(rows) + 1}"


def write_summary_sheet(ws, meta: dict, rows: list[tuple], out_name: str) -> None:
    ws.title = "测试概览"
    pass_n = sum(1 for r in rows if r[5] == "通过")
    fail_n = sum(1 for r in rows if r[5] == "失败")
    pend_n = sum(1 for r in rows if r[5] == "待验证")

    info = [
        ("技能名称", meta.get("skill_name", "")),
        ("所属武将", meta.get("hero", "")),
        ("武将势力", meta.get("faction", "")),
        ("技能分类", meta.get("category", "")),
        ("测试用例总数", len(rows)),
        ("通过", pass_n),
        ("失败", fail_n),
        ("待验证", pend_n),
        ("代码版本", meta.get("commit", "")),
        ("实现路径", meta.get("impl_path", "")),
        ("测试方式", meta.get("test_method", "")),
        ("报告文件", out_name),
    ]

    ws["A1"] = f"{meta.get('skill_name', '技能')} — 技能测试报告"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:D1")

    for i, (k, v) in enumerate(info, 3):
        ws.cell(i, 1, k).font = Font(bold=True)
        ws.cell(i, 2, str(v)).alignment = WRAP

    notes = meta.get("diff_notes") or []
    if notes:
        start = len(info) + 4
        ws.cell(start, 1, "实现与文案差异").font = Font(bold=True, size=12)
        for j, note in enumerate(notes, start + 1):
            ws.cell(j, 1, note)
            ws.merge_cells(start_row=j, start_column=1, end_row=j, end_column=4)

    fails = [f"{r[0]} {r[1]}" for r in rows if r[5] == "失败"]
    if fails:
        start = len(info) + 4 + (len(notes) + 1 if notes else 0) + 1
        ws.cell(start, 1, "失败用例").font = Font(bold=True, size=12)
        for j, f in enumerate(fails, start + 1):
            ws.cell(j, 1, f)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70


def export_xlsx(skill_dir: Path, data: dict, output: Path | None = None) -> Path:
    skill_dir = skill_dir.resolve()
    testcase_md = resolve_testcase_md(skill_dir)
    if not testcase_md.is_file():
        raise FileNotFoundError(f"未找到用例文件: {testcase_md}")

    meta = merge_report_meta(skill_dir, data)
    raw_rows = data.get("rows") or []
    rows = [row_from_obj(r) if isinstance(r, dict) else tuple(r) for r in raw_rows]
    for i, r in enumerate(rows):
        if len(r) != 7:
            raise ValueError(f"第 {i + 1} 条用例字段数应为 7，实际为 {len(r)}")
        rows[i] = (r[0], r[1], r[2], r[3], r[4], normalize_result(str(r[5])), r[6])

    out_name = f"测试报告_{meta['skill_name']}.xlsx"
    out_path = output or (skill_dir / out_name)

    wb = Workbook()
    write_summary_sheet(wb.active, meta, rows, out_name)
    detail = wb.create_sheet("测试明细")
    wb._sheets.remove(detail)
    wb._sheets.append(detail)
    write_detail_sheet(detail, rows)
    wb.save(out_path)
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(description="导出 skill-tester 测试报告为 Excel")
    parser.add_argument("--skill-dir", required=True, help="技能文件夹路径（含 testcase_*.md）")
    parser.add_argument("--input", required=True, help="报告 JSON 文件路径，或 - 表示 stdin")
    parser.add_argument("--output", help="输出 xlsx 路径（默认: 技能目录/测试报告_{技能名}.xlsx）")
    args = parser.parse_args()

    skill_dir = Path(args.skill_dir)
    if args.input == "-":
        data = json.load(sys.stdin)
    else:
        data = json.loads(Path(args.input).read_text(encoding="utf-8"))

    out = Path(args.output) if args.output else None
    path = export_xlsx(skill_dir, data, out)
    print(path)


if __name__ == "__main__":
    main()
