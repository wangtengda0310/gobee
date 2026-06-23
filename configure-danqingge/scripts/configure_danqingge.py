#!/usr/bin/env python3
"""
丹青阁配置脚本 — 自动化 Draw.xlsx / ShopGoods_商品表.xlsx / Drop.xlsx 的全链路配置。

用法见 SKILL.md 或 python configure_danqingge.py --help
"""

import argparse
import copy
import sys
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    print("错误: 需要 openpyxl 库")
    sys.exit(1)


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy.copy(src_cell.protection)
        dst_cell.alignment = copy.copy(src_cell.alignment)


def copy_row_style(src_ws, src_row, dst_ws, dst_row, max_col):
    for c in range(1, max_col + 1):
        src_cell = src_ws.cell(row=src_row, column=c)
        dst_cell = dst_ws.cell(row=dst_row, column=c)
        copy_cell_style(src_cell, dst_cell)


def safe_insert_rows(ws, start_row, count=1):
    """安全地在 start_row 位置插入 count 行，完整保留所有格式。

    openpyxl 原生的 ws.insert_rows() 存在已知 bug：
    - 合并单元格引用不会自动偏移
    - 条件格式范围不会自动调整
    - 数据验证引用不会更新
    - 行高(row_dimensions)不会跟着移动

    本函数从下往上逐行搬移数据+样式，并修正以上所有遗漏层。
    """
    import openpyxl
    from openpyxl.worksheet.dimensions import RowDimension

    max_row_before = ws.max_row
    max_col = ws.max_column

    # ── 1. 保存合并单元格 ──

    # ── 2. 从下往上逐行搬移（倒序，避免覆盖）──
    for row in range(max_row_before, start_row - 1, -1):
        new_row = row + count
        for col in range(1, max_col + 1):
            src = ws.cell(row=row, column=col)
            dst = ws.cell(row=new_row, column=col)
            dst.value = src.value
            if src.has_style:
                dst.font = copy.copy(src.font)
                dst.fill = copy.copy(src.fill)
                dst.border = copy.copy(src.border)
                dst.alignment = copy.copy(src.alignment)
                dst.number_format = src.number_format
                dst.protection = copy.copy(src.protection)

    # ── 3. 清空插入区域 ──
    for row in range(start_row, start_row + count):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None

    # ── 4. 偏移合并单元格引用（原地修改，避免 clear/re-add 导致格式丢失）──
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= start_row:
            mc.shift(0, count)

    # ── 5. 偏移行尺寸 ──
    for r in range(max_row_before, start_row - 1, -1):
        rd_old = ws.row_dimensions.get(r)
        if rd_old is not None:
            rd_new = ws.row_dimensions.get(r + count)
            if rd_new is None:
                rd_new = RowDimension(worksheet=ws, index=r + count)
                ws.row_dimensions[r + count] = rd_new
            rd_new.height = rd_old.height
            rd_new.hidden = rd_old.hidden
            try: rd_new.customFormat = rd_old.customFormat
            except AttributeError: pass
            try: rd_new.customHeight = rd_old.customHeight
            except AttributeError: pass

    return list(range(start_row, start_row + count))


def insert_row(ws, ref_row, data_dict, copy_style_from_row, max_col=None):
    """在 ref_row 下方安全插入新行，copy_style_from_row 复制样式"""
    if max_col is None:
        max_col = ws.max_column
    safe_insert_rows(ws, ref_row + 1, 1)
    new_row = ref_row + 1
    copy_row_style(ws, copy_style_from_row, ws, new_row, max_col)
    for col, value in data_dict.items():
        ws.cell(row=new_row, column=col, value=value)
    return new_row


def find_row_by_column(ws, col, value, start_row=5):
    for r in range(start_row, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=col).value
        if cell_val is not None and str(cell_val) == str(value):
            return r
    return None


def find_all_rows_by_column(ws, col, value, start_row=5):
    """返回列中所有匹配值的行号列表（从上到下）。"""
    found = []
    for r in range(start_row, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=col).value
        if cell_val is not None and str(cell_val) == str(value):
            found.append(r)
    return found


def find_last_in_group(ws, group_col, group_value, start_row=5):
    last_row = None
    for r in range(start_row, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=group_col).value
        if cell_val is not None and str(cell_val) == str(group_value):
            last_row = r
    return last_row


def build_header_map(ws, header_row=3):
    """读取表头行，返回 {字段名: 1-based列号} 的映射。
    优先读 Row 3（英文字段名），为空退至 Row 2。"""
    hdr = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            hdr[str(cell.value).strip()] = cell.column
    if not hdr:
        for cell in ws[header_row - 1]:
            if cell.value is not None:
                hdr[str(cell.value).strip()] = cell.column
    return hdr


# ============ Step 6: DrawSkin 配置 ============

def configure_draw_skin(config_dir, args):
    path = f"{config_dir}/Draw.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb['皮肤抽奖|DrawSkin']
    max_col = ws.max_column

    # 动态表头映射（硬编码值作为 fallback）
    hdr = build_header_map(ws)
    c_id = hdr.get('Id', 1)
    c_name = hdr.get('Name', 2)
    c_bigitem = hdr.get('BigAwardItemId', 13)
    c_start = hdr.get('StartTime', 14)
    c_end = hdr.get('EndTime', 15)
    c_bg = hdr.get('BgSpinePath', 22)
    c_skinname = hdr.get('SkinNameSpriteName', 23)
    c_byproduct = hdr.get('byproduct', 24)
    c_titleicon = hdr.get('TitleIconName', 29)

    draw_id = 2000 + args.period
    prev_draw_id = draw_id - 1

    # 查找所有 N-1 期行 → 避免定位到下方的预设占位行
    all_prev_rows = find_all_rows_by_column(ws, c_id, prev_draw_id)
    if not all_prev_rows:
        print(f"错误: 在 DrawSkin 中找不到 Id={prev_draw_id}")
        wb.close()
        return None

    if len(all_prev_rows) >= 2:
        print(f"  ⚠️ Id={prev_draw_id} 匹配到 {len(all_prev_rows)} 行: {all_prev_rows}")
        print(f"  📍 使用第 1 行 Row {all_prev_rows[0]} 作为插入参考（忽略下方预设占位行）")
    ref_row = all_prev_rows[0]  # 始终取第一个（顶部实际数据行）

    # BgSpinePath 替换
    old_bg = str(ws.cell(row=ref_row, column=c_bg).value or '')
    new_bg = old_bg.replace(f"hero_{args.prev_hero_pinyin}_{args.prev_skin_pinyin}",
                            f"hero_{args.hero_pinyin}_{args.skin_pinyin}")

    # TitleIconName 替换
    old_ti = str(ws.cell(row=ref_row, column=c_titleicon).value or '')
    new_ti = old_ti.replace(f"_{args.prev_skin_pinyin}", f"_{args.skin_pinyin}")

    byproduct = f"{args.skin_item},{args.mengjiang_item},{args.byproduct_item}"

    # 从 ref_row 全量复制所有列作为基值
    data = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=ref_row, column=c).value
        if v is not None:
            data[c] = v

    # 只覆盖差异字段
    data[c_id] = draw_id
    data[c_name] = args.skin_name
    data[c_bigitem] = int(args.skin_item)
    data[c_start] = args.start_time
    data[c_end] = args.end_time
    data[c_bg] = new_bg
    data[c_skinname] = args.skin_name
    data[c_byproduct] = byproduct
    data[c_titleicon] = new_ti

    new_row = insert_row(ws, ref_row, data, ref_row, max_col)

    # 检查并删除同一 Id 的预设占位行（Note #13）
    for r in range(ref_row + 1, ws.max_row + 1):
        if r == new_row:
            continue
        if ws.cell(row=r, column=c_id).value == draw_id:
            ws.delete_rows(r)
            print(f"  🗑️ 删除预设占位行 Row {r} (Id={draw_id})")
            break

    wb.save(path)
    wb.close()
    print(f"✅ Draw.xlsx → DrawSkin Row {new_row} (Id={draw_id}, {args.skin_name})")
    return new_row


# ============ Step 8: ShopGoods 配置 ============

def configure_shop_goods(config_dir, args):
    """
    ShopGoods 列映射 (1-based):
    Col  1(A): 商品id (Id)
    Col  2(B): 商品名称 (Name)
    Col  3(C): 商品说明 (Desc)
    Col  4(D): 获得道具 (Item, 格式: {id;1})
    Col  6(F): 购买消耗货币ID (CostId)
    Col  7(G): 商店类型 (ShopType)
    Col 12(L): 价格 (Price)
    Col 13(M): 原价 (OldPrice)
    Col 22(V): 商品图标 (Icon)
    Col 23(W): 上架时间 (OnShelfTime)
    Col 24(X): 下架时间 (OffShelfTime) — 固定 2060-10-31
    Col 33(AG): 商品在购买界面图标 (IconInBuyWindow)
    Col 35(AI): 奖励预览窗口 (RewardID, 道具id)
    Col 36(AJ): 商品显示页签 (ItemDisplayTap): 0=皮肤,3=萌将,1=边框,2=手牌皮肤
    """
    path = f"{config_dir}/ShopGoods_商品表.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb['商品表|ShopGood']
    hdr = build_header_map(ws)

    c_id = hdr.get('Id', 1)
    c_name = hdr.get('Name', 2)
    c_desc = hdr.get('Desc', 3)
    c_item = hdr.get('Item', 4)
    c_icon = hdr.get('Icon', 23)
    c_onshelf = hdr.get('OnShelfTime', 24)
    c_offshelf = hdr.get('OffShelfTime', 25)
    c_buyicon = hdr.get('IconInBuyWindow', 34)
    c_reward = hdr.get('RewardID', 36)
    c_tab = hdr.get('ItemDisplayTap', 37)

    # 找 #丹青 标记行
    dq_marker_row = find_row_by_column(ws, c_id, '#丹青')
    if dq_marker_row is None:
        print("错误: 在 ShopGoods 中找不到 #丹青 标记")
        wb.close()
        return None

    # 找 #公会 标记行 → 插入位置 = #公会 前一行
    gh_marker_row = find_row_by_column(ws, c_id, '#公会')
    if gh_marker_row is None:
        print("错误: 在 ShopGoods 中找不到 #公会 标记")
        wb.close()
        return None

    # 插入位置: #公会 上方 (即当前最后一条丹青数据行下方)
    insert_pos = gh_marker_row - 1
    ref_row = dq_marker_row + 1  # 丹青区域第一行作为样式参考
    max_col = ws.max_column

    def get_next_id():
        prev_id = int(ws.cell(row=insert_pos, column=c_id).value or 0)
        return prev_id + 1

    def fill_defaults(out_dict):
        """从 ref_row 复制所有未设置的列"""
        for c in range(1, max_col + 1):
            if c not in out_dict:
                val = ws.cell(row=ref_row, column=c).value
                if val is not None:
                    out_dict[c] = val

    # --- 8a: 皮肤商品 (Tab=0) ---
    skin_desc = (
        f"购买后获得{args.prev_hero_name}的武将皮肤，"
        f"获得后可以在游戏中改变武将{args.prev_hero_name}的卡牌形象"
    )
    skin_data = {
        c_id: get_next_id(),
        c_name: args.prev_skin_name,
        c_desc: skin_desc,
        c_item: f"{{{args.prev_skin_item};1}}",  # 输出: {道具id;1} (花括号是字面量)
        c_icon: args.prev_skin_icon,
        c_onshelf: args.start_time,
        c_offshelf: '2060-10-31 23:59:59',
        c_buyicon: args.prev_skin_icon,
        c_reward: int(args.prev_skin_item),
        c_tab: 0,  # 皮肤=0
    }
    fill_defaults(skin_data)

    skin_row = insert_row(ws, insert_pos, skin_data, ref_row, max_col)
    insert_pos += 1
    print(f"✅ ShopGoods Row {skin_row} (皮肤: {args.prev_skin_name}, Tab=0)")

    # --- 8b: 萌将商品 (Tab=3) ---
    mj_display_name = f"{args.prev_hero_name}·萌将"
    mj_desc = (
        f"购买后获得{args.prev_hero_name}·萌将形象，"
        f"可以在形象系统中使用，将自己的形象设置为{args.prev_hero_name}·萌将形象"
    )
    mj_data = {
        c_id: get_next_id(),
        c_name: mj_display_name,
        c_desc: mj_desc,
        c_item: f"{{{args.prev_mengjiang_item};1}}",  # 输出: {道具id;1} (花括号是字面量)
        c_icon: args.prev_mengjiang_icon,
        c_onshelf: args.start_time,
        c_offshelf: '2060-10-31 23:59:59',
        c_buyicon: args.prev_mengjiang_icon,
        c_reward: int(args.prev_mengjiang_item),
        c_tab: 3,  # 萌将=3
    }
    fill_defaults(mj_data)

    mj_row = insert_row(ws, insert_pos, mj_data, ref_row, max_col)
    insert_pos += 1
    print(f"✅ ShopGoods Row {mj_row} (萌将: {mj_display_name}, Tab=3)")

    # --- 8c: 副产物商品 ---
    bp_type = args.prev_byproduct_type
    if bp_type == 'frame':
        bp_tab = 1  # 边框=1
        bp_desc = (
            f"购买后获得{args.prev_byproduct_name}形象边框，"
            f"个人形象边框道具，可以在形象系统中使用，将自己的形象边框设置为此边框形象"
        )
    else:
        bp_tab = 2  # 手牌皮肤=2
        bp_desc = (
            f"购买后获得{args.prev_byproduct_name}手牌皮肤，"
            f"可以在收藏-手牌中使用，改变游戏中的手牌样式"
        )

    bp_data = {
        c_id: get_next_id(),
        c_name: args.prev_byproduct_name,
        c_desc: bp_desc,
        c_item: f"{{{args.prev_byproduct_item};1}}",  # 输出: {道具id;1} (花括号是字面量)
        c_icon: args.prev_byproduct_icon,
        c_onshelf: args.start_time,
        c_offshelf: '2060-10-31 23:59:59',
        c_buyicon: args.prev_byproduct_icon,
        c_reward: int(args.prev_byproduct_item),
        c_tab: bp_tab,
    }
    fill_defaults(bp_data)

    bp_row = insert_row(ws, insert_pos, bp_data, ref_row, max_col)
    print(f"✅ ShopGoods Row {bp_row} (副产物: {args.prev_byproduct_name}, Tab={bp_tab})")

    wb.save(path)
    wb.close()


# ============ Step 9: DropItem 配置 ============

def find_first_fixed_in_group(ws, group_col, group_value, start_row=5):
    """找到分组中第一个固定奖励行（ExpireDate 含 2055/2054 等远期时间）"""
    for r in range(start_row, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=group_col).value
        if cell_val is not None and str(cell_val) == str(group_value):
            end = str(ws.cell(row=r, column=13).value or '')
            if '2055' in end or '2054' in end:
                return r
    return None


def configure_drop(config_dir, args):
    """DropItem 列映射 (1-based):
    Col  1(A): Id
    Col  2(B): Name
    Col  3(C): DropGroup
    Col  4(D): Item (格式: {id;1})
    ...
    Col 12(L): ValidDate
    Col 13(M): ExpireDate

    插入规则: 组内有固定奖励行(ExpireDate 含 2054/2055)时，在其上方插入；
    否则在组末尾插入。
    """

    path = f"{config_dir}/Drop.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb['掉落道具表|DropItem']
    max_col = ws.max_column
    hdr = build_header_map(ws)

    c_id = hdr.get('Id', 1)
    c_name = hdr.get('Name', 2)
    c_group = hdr.get('DropGroup', 3)
    c_item = hdr.get('Item', 4)
    c_valid = hdr.get('ValidDate', 12)
    c_expire = hdr.get('ExpireDate', 13)

    groups_config = [
        (90001, args.skin_name, args.skin_item, "皮肤"),
        (90002, args.skin_name, args.skin_item, "皮肤(保底)"),
        (90003, f"{args.hero_name}·萌将", args.mengjiang_item, "萌将"),
        (90005, args.byproduct_name, args.byproduct_item, "副产物"),
    ]

    for group_id, name, item_id, label in groups_config:
        first_fixed = find_first_fixed_in_group(ws, c_group, group_id)
        if first_fixed is not None:
            # 有固定奖励行 → 在其上方安全插入空行，再填充数据
            safe_insert_rows(ws, first_fixed, 1)
            new_row = first_fixed
            ref_row = new_row - 1  # 上一期皮肤行（在固定奖励上方）
            copy_row_style(ws, ref_row, ws, new_row, max_col)
            print(f"  Group {group_id}: 在固定奖励行上方插入 (Row {new_row})")
        else:
            # 无固定奖励行 → 在组末尾插入
            last_row = find_last_in_group(ws, c_group, group_id)
            if last_row is None:
                print(f"警告: 找不到 DropGroup={group_id} 的数据行，跳过")
                continue
            ref_row = last_row
            new_row = insert_row(ws, ref_row, {}, ref_row, max_col)
            print(f"  Group {group_id}: 在组末尾插入 (Row {new_row})")

        drop_data = {
            c_id: int(ws.cell(row=ref_row, column=c_id).value or 0) + 1,
            c_name: name,
            c_group: group_id,
            c_item: f"{{{item_id};1}}",  # 输出: {道具id;1} (花括号是字面量)
            c_valid: args.start_time,
            c_expire: args.end_time,
        }

        # 从 ref_row 复制其他列
        for c in range(1, max_col + 1):
            if c not in drop_data:
                val = ws.cell(row=ref_row, column=c).value
                if val is not None:
                    drop_data[c] = val

        for col, value in drop_data.items():
            ws.cell(row=new_row, column=col, value=value)

        print(f"✅ DropItem: Group={group_id}, Row {new_row} ({label}: {name})")

    wb.save(path)
    wb.close()


# ============ Main ============

def main():
    parser = argparse.ArgumentParser(description='配置丹青阁')

    parser.add_argument('config_dir', help='配置文件目录 (如 /home/.../名将杀配置/)')
    parser.add_argument('period', type=int, help='丹青阁期数 (如 21)')

    # 当期 (N期)
    parser.add_argument('--skin-name', required=True, help='皮肤名称')
    parser.add_argument('--skin-pinyin', required=True, help='皮肤拼音')
    parser.add_argument('--skin-item', required=True, help='皮肤道具id')
    parser.add_argument('--skin-icon', default='待配置', help='皮肤图标')

    parser.add_argument('--hero-name', required=True, help='武将名')
    parser.add_argument('--hero-pinyin', required=True, help='武将拼音')

    parser.add_argument('--mengjiang-item', required=True, help='萌将道具id')
    parser.add_argument('--mengjiang-icon', default='待配置', help='萌将图标')

    parser.add_argument('--byproduct-name', required=True, help='副产物名称')
    parser.add_argument('--byproduct-item', required=True, help='副产物道具id')
    parser.add_argument('--byproduct-icon', default='待配置', help='副产物图标')
    parser.add_argument('--byproduct-type', choices=['frame', 'cardskin'],
                        help='副产物类型: frame(边框) / cardskin(手牌皮肤)')

    # 时间
    parser.add_argument('--start-time', required=True, help='开始时间 (YYYY-MM-DD HH:MM:SS)')
    parser.add_argument('--end-time', required=True, help='结束时间 (YYYY-MM-DD HH:MM:SS)')

    # 上一期 (N-1期) — 用于 ShopGoods
    parser.add_argument('--prev-skin-name', required=True, help='上一期皮肤名称')
    parser.add_argument('--prev-skin-item', required=True, help='上一期皮肤道具id')
    parser.add_argument('--prev-skin-icon', default='待配置', help='上一期皮肤图标')
    parser.add_argument('--prev-skin-pinyin', help='上一期皮肤拼音 (BgSpinePath替换用)')

    parser.add_argument('--prev-hero-name', required=True, help='上一期武将名')
    parser.add_argument('--prev-hero-pinyin', help='上一期武将拼音 (BgSpinePath替换用)')

    parser.add_argument('--prev-mengjiang-name', required=True, help='上一期萌将名')
    parser.add_argument('--prev-mengjiang-item', required=True, help='上一期萌将道具id')
    parser.add_argument('--prev-mengjiang-icon', default='待配置', help='上一期萌将图标')

    parser.add_argument('--prev-byproduct-name', required=True, help='上一期副产物名称')
    parser.add_argument('--prev-byproduct-item', required=True, help='上一期副产物道具id')
    parser.add_argument('--prev-byproduct-icon', default='待配置', help='上一期副产物图标')
    parser.add_argument('--prev-byproduct-type', choices=['frame', 'cardskin'],
                        help='上一期副产物类型')

    args = parser.parse_args()

    print(f"{'='*60}")
    print(f"配置丹青阁-{args.period}期")
    print(f"  武将: {args.hero_name} ({args.hero_pinyin})")
    print(f"  皮肤: {args.skin_name} (ID={args.skin_item})")
    print(f"  萌将: ID={args.mengjiang_item}")
    print(f"  副产物: {args.byproduct_name} (ID={args.byproduct_item}, 类型={args.byproduct_type})")
    print(f"  时间: {args.start_time} ~ {args.end_time}")
    print(f"  上一期皮肤: {args.prev_skin_name} (ID={args.prev_skin_item})")
    print(f"{'='*60}")

    # Step 6: Draw.xlsx
    configure_draw_skin(args.config_dir, args)
    print()

    # Step 8: ShopGoods
    configure_shop_goods(args.config_dir, args)
    print()

    # Step 9: Drop.xlsx
    configure_drop(args.config_dir, args)
    print()

    print(f"{'='*60}")
    print("🎉 丹青阁配置完成!")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
