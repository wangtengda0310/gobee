#!/usr/bin/env python3
"""
名将杀 - 武将掉落时间配置脚本
用法: python3 configure_hero_drop.py <武将名>
示例: python3 configure_hero_drop.py 郑国

流程:
1. 从飞书 Wiki "武将排班表" 读取武将信息（发布时间/品质/投放方式）
2. 从 Item.xlsx 查找武将道具 ID
3. 在 Drop.xlsx 的掉落道具表|DropItem 中插入新掉落配置行
"""

import sys
import subprocess
import json
import os
from datetime import datetime, timedelta
from copy import copy

# === 配置常量 ===
PYTHON_VENV = os.path.expanduser(
    "~/.openclaw/skills/aicconfig/scripts/aicconfig/venv/bin/python3"
)
EXCEL_DIR = os.path.expanduser("~/.openclaw/workspace/名将杀配置")

# 飞书 Wiki 配置
WIKI_TOKEN = "XwMDwfskviuqL1kCtDQc7hEEnxe"
SPREADSHEET_TOKEN = "HODRsCYqVhuLGttKm61csaDAngf"
SHEET_ID = "d50c43"

# 品质 → 掉落组ID 映射
QUALITY_DROP_GROUP = {
    "传说": 10004,
    "史诗": 10003,
    "稀有": 10002,
    "普通": 10001,
}

# 投放方式 → 自然月偏移
DEPLOY_MONTH_OFFSET = {
    "大将军": 1,
    "战令": 4,
}

# Excel 日期基准 (1900 系统)
EXCEL_DATE_BASE = datetime(1899, 12, 30)


def excel_date_to_str(excel_val):
    """将 Excel 整数日期转换为 YYYY-MM-DD 字符串"""
    try:
        days = int(excel_val)
        dt = EXCEL_DATE_BASE + timedelta(days=days)
        return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def add_months(dt, months):
    """给日期增加自然月"""
    year = dt.year
    month = dt.month + months
    while month > 12:
        year += 1
        month -= 12
    # 处理月末溢出（如 1月31日 + 1月 → 2月28/29日）
    import calendar
    max_day = calendar.monthrange(year, month)[1]
    day = min(dt.day, max_day)
    return datetime(year, month, day, dt.hour, dt.minute, dt.second)


def lark_cli(*args):
    """调用 lark-cli 并返回 JSON"""
    cmd = ["lark-cli"] + list(args)
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
    if result.returncode != 0:
        print(f"lark-cli 错误: {result.stderr}")
        return None
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError:
        print(f"lark-cli 返回非 JSON: {result.stdout[:200]}")
        return None


def find_hero_in_wiki(hero_name):
    """Step 1: 从飞书 Wiki 查找武将信息"""
    print(f"🔍 在飞书 Wiki 中搜索武将: {hero_name}")

    # 使用 find 定位武将所在行
    result = lark_cli(
        "sheets", "+find",
        "--spreadsheet-token", SPREADSHEET_TOKEN,
        "--sheet-id", SHEET_ID,
        "--find", hero_name,
    )
    if not result or not result.get("ok"):
        print("❌ 搜索武将失败")
        return None

    cells = result.get("data", {}).get("find_result", {}).get("matched_cells", [])
    if not cells:
        print(f"❌ 未找到武将: {hero_name}")
        return None

    # 提取行号
    cell_ref = cells[0]  # e.g. "B146"
    row_num = int(cell_ref[1:])

    # 读取该行数据 (列 A-O)
    range_str = f"{SHEET_ID}!A{row_num}:O{row_num}"
    result = lark_cli(
        "sheets", "+read",
        "--spreadsheet-token", SPREADSHEET_TOKEN,
        "--sheet-id", SHEET_ID,
        "--range", range_str,
    )
    if not result or not result.get("ok"):
        print("❌ 读取武将行数据失败")
        return None

    values = result.get("data", {}).get("valueRange", {}).get("values", [[]])
    row_data = values[0] if values else []

    # 列映射: A-时代 B-武将 C-类型 D-开发完成时间 E-发布时间 F-状态 G-AI H-武将成就
    #          I-草图 J-细化 K-spine L-武将信物 M-品质 N-投放方式 O-投放时间
    def safe_get(idx):
        return row_data[idx] if idx < len(row_data) else None

    release_date_num = safe_get(4)    # E - 发布时间
    quality = safe_get(12)             # M - 品质
    deploy_method = safe_get(13)       # N - 投放方式

    release_date = excel_date_to_str(release_date_num) if release_date_num else None

    info = {
        "hero_name": hero_name,
        "row": row_num,
        "release_date_num": release_date_num,
        "release_date": release_date,
        "quality": quality,
        "deploy_method": deploy_method,
    }

    print(f"  📋 武将信息:")
    print(f"     发布时间(原始): {release_date_num} → {release_date}")
    print(f"     品质: {quality}")
    print(f"     投放方式: {deploy_method if deploy_method else '(空)'}")

    return info


def find_hero_item_id(hero_name):
    """Step 2: 从 Item.xlsx 查找武将道具 ID (Type=Hero)"""
    import openpyxl
    print(f"🔍 在 Item.xlsx 中搜索武将道具: {hero_name}")

    item_path = os.path.join(EXCEL_DIR, "Item.xlsx")
    wb = openpyxl.load_workbook(item_path, data_only=True)
    ws = wb["道具表|Item"]

    for row_idx in range(5, ws.max_row + 1):
        name_val = ws.cell(row_idx, 2).value   # B - Name
        type_val = ws.cell(row_idx, 3).value    # C - Type
        if name_val and str(name_val).strip() == hero_name:
            if type_val and str(type_val).strip() == "Hero":
                item_id = ws.cell(row_idx, 1).value  # A - Id
                wb.close()
                print(f"  ✅ 找到道具: Id={item_id}, Row={row_idx}")
                return int(item_id)

    wb.close()
    print(f"  ❌ 未找到武将 {hero_name} 的 Hero 类型道具")
    return None


def calc_drop_time(release_date_str, deploy_method):
    """计算掉落时间"""
    if not release_date_str:
        return None

    release_dt = datetime.strptime(release_date_str, "%Y-%m-%d")

    if not deploy_method or str(deploy_method).strip() == "":
        # 投放方式为空 → 发布时间 05:00:00
        return f"{release_date_str} 05:00:00"

    months = DEPLOY_MONTH_OFFSET.get(deploy_method)
    if months:
        # 发布时间 + N个自然月
        target_dt = add_months(release_dt, months)
        return target_dt.strftime("%Y-%m-%d 00:00:00")

    # fallback
    return f"{release_date_str} 05:00:00"


def configure_drop(hero_info, item_id):
    """Step 3-8: 在 Drop.xlsx 中插入新掉落配置行"""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    hero_name = hero_info["hero_name"]
    quality = hero_info["quality"]
    deploy_method = hero_info["deploy_method"]
    release_date = hero_info["release_date"]

    # 掉落组ID
    drop_group = QUALITY_DROP_GROUP.get(quality)
    if not drop_group:
        print(f"❌ 未知品质: {quality}")
        return False

    # 掉落时间
    drop_time = calc_drop_time(release_date, deploy_method)
    print(f"  ⏰ 掉落时间: {drop_time}")

    # 掉落道具格式
    drop_item_str = f"{{{item_id};1}}"

    # 打开 Drop.xlsx
    drop_path = os.path.join(EXCEL_DIR, "Drop.xlsx")
    print(f"📂 打开 Drop.xlsx: {drop_path}")
    wb = openpyxl.load_workbook(drop_path)
    ws = wb["掉落道具表|DropItem"]

    # Step 3: 找到 #武将掉落 区域
    hero_section_start = None
    hero_section_end = None

    for row_idx in range(1, ws.max_row + 1):
        cell_val = str(ws.cell(row_idx, 1).value or "")
        if cell_val.startswith("#"):
            if "武将掉落" in cell_val:
                hero_section_start = row_idx
            elif hero_section_start and not hero_section_end:
                hero_section_end = row_idx - 1
                break

    if not hero_section_start:
        print("❌ 未找到 #武将掉落 区域")
        wb.close()
        return False

    if not hero_section_end:
        hero_section_end = ws.max_row

    print(f"  📍 #武将掉落 区域: Row {hero_section_start+1} ~ Row {hero_section_end}")

    # Step 4: 找到最大掉落ID
    max_id = 0
    last_data_row = hero_section_start + 1
    for row_idx in range(hero_section_start + 1, hero_section_end + 1):
        val = ws.cell(row_idx, 1).value
        if val is not None and str(val).strip():
            try:
                id_val = int(val)
                if id_val > max_id:
                    max_id = id_val
                    last_data_row = row_idx
            except (ValueError, TypeError):
                pass

    new_id = max_id + 1
    print(f"  📊 当前最大掉落ID: {max_id}, 新掉落ID: {new_id}")
    print(f"  📍 最后数据行: Row {last_data_row}")

    # Step 5: 在最后一行下方插入新行
    insert_row = last_data_row + 1
    ws.insert_rows(insert_row)
    print(f"  ➕ 在 Row {insert_row} 插入新行")

    # Step 6: 复制上一行样式
    src_row = insert_row - 1  # 插入后，原来的最后行现在在 insert_row-1

    # 如果插入位置越过 section end，需要调整
    if insert_row <= hero_section_end:
        hero_section_end += 1

    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(src_row, col)
        dst_cell = ws.cell(insert_row, col)

        # 复制值和样式
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)
            dst_cell.alignment = copy(src_cell.alignment)

        # 复制值作为默认值（会被覆盖的列除外）
        if col not in [1, 2, 3, 4, 12]:  # Id, Name, DropGroup, Item, ValidDate
            dst_cell.value = src_cell.value

    # Step 7: 写入新值
    ws.cell(insert_row, 1).value = new_id      # Id
    ws.cell(insert_row, 2).value = hero_name    # Name
    ws.cell(insert_row, 3).value = drop_group    # DropGroup
    ws.cell(insert_row, 4).value = drop_item_str  # Item
    ws.cell(insert_row, 12).value = drop_time    # ValidDate
    ws.cell(insert_row, 13).value = "2054-12-31 00:00:00"  # ExpireDate

    print(f"  ✏️ 写入配置:")
    print(f"     Id={new_id}, Name={hero_name}")
    print(f"     DropGroup={drop_group}, Item={drop_item_str}")
    print(f"     ValidDate={drop_time}")

    # 保存
    wb.save(drop_path)
    wb.close()
    print(f"  ✅ Drop.xlsx 保存成功!")

    return True


def main():
    if len(sys.argv) < 2:
        print("用法: python3 configure_hero_drop.py <武将名>")
        print("示例: python3 configure_hero_drop.py 郑国")
        sys.exit(1)

    hero_name = sys.argv[1]
    print(f"🐙 开始配置武将掉落时间: {hero_name}")
    print("=" * 50)

    # Step 1: 从 Wiki 获取武将信息
    hero_info = find_hero_in_wiki(hero_name)
    if not hero_info:
        sys.exit(1)

    # Step 2: 从 Item.xlsx 获取道具 ID
    item_id = find_hero_item_id(hero_name)
    if not item_id:
        sys.exit(1)

    # Step 3-8: 配置 Drop.xlsx
    success = configure_drop(hero_info, item_id)
    if success:
        print(f"\n🎉 完成! 武将 {hero_name} 的掉落时间已配置完成")
        print(f"   道具ID: {item_id}")
        print(f"   掉落组ID: {QUALITY_DROP_GROUP.get(hero_info['quality'])} ({hero_info['quality']})")
        print(f"   发布: {hero_info['release_date']}")
        print(f"   投放方式: {hero_info['deploy_method'] if hero_info['deploy_method'] else '(空)'}")
        print(f"   掉落时间: {calc_drop_time(hero_info['release_date'], hero_info['deploy_method'])}")
    else:
        print("\n❌ 配置失败")
        sys.exit(1)


if __name__ == "__main__":
    main()
