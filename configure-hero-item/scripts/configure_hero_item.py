#!/usr/bin/env python3
"""configure_hero_item.py - Configure a new hero item across all Excel config files."""

import argparse, os, sys
from copy import copy
import openpyxl

QUALITY_MAP = {
    "传说": {"rarity": 4, "resolve": "{1000025;400}", "synthetic": "{1000025;1600}"},
    "史诗": {"rarity": 3, "resolve": "{1000025;100}", "synthetic": "{1000025;400}"},
    "稀有": {"rarity": 2, "resolve": "{1000025;20}",  "synthetic": "{1000025;80}"},
    "普通": {"rarity": 1, "resolve": "{1000025;5}",   "synthetic": "{1000025;20}"},
}

def copy_cell_style(src, dst):
    dst.value = src.value
    if src.has_style:
        dst.font = copy(src.font); dst.border = copy(src.border)
        dst.fill = copy(src.fill); dst.number_format = copy(src.number_format)
        dst.protection = copy(src.protection); dst.alignment = copy(src.alignment)

def copy_row_style(ws, src_row, dst_row, max_col):
    for col in range(1, max_col + 1):
        copy_cell_style(ws.cell(row=src_row, column=col), ws.cell(row=dst_row, column=col))

def safe_insert_rows(ws, start_row, count=1):
    """安全地在 start_row 位置插入 count 行，完整保留所有格式。

    openpyxl 原生的 ws.insert_rows() 存在已知 bug：
    - 合并单元格引用不会自动偏移
    - 条件格式范围不会自动调整
    - 数据验证引用不会更新
    - 行高(row_dimensions)不会跟着移动

    本函数从下往上逐行搬移数据+样式，并修正以上所有遗漏层。
    """
    from openpyxl.worksheet.dimensions import RowDimension

    max_row_before = ws.max_row
    max_col = ws.max_column

    old_merges = [str(mc) for mc in ws.merged_cells.ranges]

    for row in range(max_row_before, start_row - 1, -1):
        new_row = row + count
        for col in range(1, max_col + 1):
            src = ws.cell(row=row, column=col)
            dst = ws.cell(row=new_row, column=col)
            dst.value = src.value
            if src.has_style:
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
                dst.protection = copy(src.protection)

    for row in range(start_row, start_row + count):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None

    ws.merged_cells.ranges.clear()
    for mc_str in old_merges:
        mc = CellRange(mc_str)
        if mc.min_row >= start_row:
            mc.shift(0, count)
        ws.merge_cells(str(mc))

    for r in range(max_row_before, start_row - 1, -1):
        rd_old = ws.row_dimensions.get(r)
        if rd_old is not None:
            rd_new = ws.row_dimensions.get(r + count)
            if rd_new is None:
                rd_new = RowDimension(worksheet=ws, index=r + count)
                ws.row_dimensions[r + count] = rd_new
            rd_new.height = rd_old.height
            rd_new.hidden = rd_old.hidden
            try: rd_new.customFormat = rd_old.customFormat
            except AttributeError: pass
            try: rd_new.customHeight = rd_old.customHeight
            except AttributeError: pass

    return list(range(start_row, start_row + count))


def insert_row_below(ws, ref_row, max_col):
    """在 ref_row 下方安全插入新行，复制样式"""
    new_row = ref_row + 1
    safe_insert_rows(ws, new_row, 1)
    copy_row_style(ws, ref_row, new_row, max_col)
    return new_row

def find_hero_data(ws_hero, hero_name):
    for row in range(5, ws_hero.max_row + 1):
        if ws_hero.cell(row=row, column=2).value is not None and str(ws_hero.cell(row=row, column=2).value).strip() == hero_name:
            return (ws_hero.cell(row=row, column=1).value,
                    ws_hero.cell(row=row, column=3).value,
                    ws_hero.cell(row=row, column=22).value)
    return None, None, None

def find_hero_lines(ws_lines, hero_name):
    start_marker = f"#{hero_name}"
    in_section = False
    debut_ids, kill_ids, dead_ids, opt_ids, skill_ids = [], [], [], [], []
    for row in range(5, ws_lines.max_row + 1):
        c1 = ws_lines.cell(row=row, column=1).value
        c1_str = str(c1).strip() if c1 is not None else ""
        if c1_str == start_marker:
            in_section = True; continue
        elif c1_str.startswith("#") and in_section:
            break
        if not in_section: continue
        line_id = ws_lines.cell(row=row, column=1).value
        if line_id is None: continue
        tag = ws_lines.cell(row=row, column=4).value
        tag_str = str(tag).strip() if tag else ""
        lid = int(line_id)
        if tag_str == "登场": debut_ids.append(lid)
        elif tag_str == "击杀": kill_ids.append(lid)
        elif tag_str == "阵亡": dead_ids.append(lid)
        elif tag_str == "自选": opt_ids.append(lid)
        else: skill_ids.append(lid)
    for lst in [debut_ids, kill_ids, dead_ids, opt_ids, skill_ids]: lst.sort()
    all_ids = sorted(skill_ids + debut_ids + kill_ids + dead_ids)
    return {
        "all_lines": ",".join(str(x) for x in all_ids) if all_ids else "",
        "debut_lines": ",".join(str(x) for x in debut_ids) if debut_ids else "",
        "kill_lines": ",".join(str(x) for x in kill_ids) if kill_ids else "",
        "dead_lines": ",".join(str(x) for x in dead_ids) if dead_ids else "",
        "opt_lines": ",".join(str(x) for x in opt_ids) if opt_ids else "",
        "found": len(all_ids) + len(opt_ids) > 0,
    }

def set_text_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    cell.number_format = '@'
    cell.value = str(value) if value else ""

# ---- STEP 1 ----
def step1_read_hero(base_dir, hero_name):
    hero_path = os.path.join(base_dir, "Hero.xlsx")
    wb = openpyxl.load_workbook(hero_path)
    ws = wb["武将表|Hero"]
    hero_id, hero_pinyin, expansion_pack = find_hero_data(ws, hero_name)
    wb.close()
    if hero_id is None:
        print(f"ERROR: Hero '{hero_name}' not found in Hero.xlsx"); sys.exit(1)
    hero_pinyin_lower = str(hero_pinyin).lower() if hero_pinyin else ""
    hero_item_id = int("10" + str(hero_id))
    print(f"\n[Step 1 - Hero.xlsx] {hero_name}: ID={hero_id}, Pinyin={hero_pinyin}, ItemId={hero_item_id}, Exp={expansion_pack}")
    return hero_id, hero_pinyin, hero_pinyin_lower, hero_item_id, expansion_pack

# ---- STEP 2 ----
def step2_item_hero(base_dir, hero_name, hero_id, hero_item_id, hero_pinyin_lower, quality_info):
    item_path = os.path.join(base_dir, "Item.xlsx")
    wb = openpyxl.load_workbook(item_path)
    ws = wb["道具表|Item"]
    max_col = ws.max_column
    section_row = None
    for row in range(5, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is not None and str(ws.cell(row=row, column=1).value).strip() == "#武将道具":
            section_row = row; break
    boundaries = ["#武将形象道具","#皮肤伴生形象","#形象边框","#卡牌皮肤","#牌局背景皮肤","#牌局UI装扮","#武将皮肤道具","#武将信物"]
    boundary_row = ws.max_row
    for row in range(section_row + 1, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1).value
        if c1 is not None and str(c1).strip() in boundaries:
            boundary_row = row - 1; break
    def to_int(v):
        try: return int(v)
        except: return None
    ref_row = section_row
    for row in range(section_row + 1, boundary_row + 1):
        val = to_int(ws.cell(row=row, column=1).value)
        if val is not None and val < hero_item_id:
            ref_row = row
    new_row = insert_row_below(ws, ref_row, max_col)
    print(f"\n[Step 2 - Item: 武将道具] Ref={ref_row} -> New={new_row}")
    ws.cell(row=new_row, column=1).value = hero_item_id
    ws.cell(row=new_row, column=2).value = hero_name
    ws.cell(row=new_row, column=4).value = quality_info["rarity"]
    ws.cell(row=new_row, column=12).value = hero_id
    ws.cell(row=new_row, column=14).value = quality_info["resolve"]
    ws.cell(row=new_row, column=15).value = int("11" + str(hero_id))
    ws.cell(row=new_row, column=22).value = f"UI/Images/Item/ui_s1_daoju_wujiang_{hero_pinyin_lower}.png"
    ws.cell(row=new_row, column=28).value = f"{hero_name}的武将卡牌，在所有游戏模式中使用武将{hero_name}"
    ws.cell(row=new_row, column=30).value = 1
    ws.cell(row=new_row, column=31).value = quality_info["synthetic"]
    print(f"  id={hero_item_id} name={hero_name} rarity={quality_info['rarity']}")
    wb.save(item_path); wb.close()
    print(f"  OK Item.xlsx (hero item)")

# ---- STEP 3 ----
def step3_item_skins(base_dir, hero_name, hero_id, hero_item_id, hero_pinyin_lower):
    item_path = os.path.join(base_dir, "Item.xlsx")
    wb = openpyxl.load_workbook(item_path)
    ws = wb["道具表|Item"]
    max_col = ws.max_column
    section_row = None
    for row in range(5, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is not None and str(ws.cell(row=row, column=1).value).strip() == "#武将皮肤道具":
            section_row = row; break
    boundary_row = ws.max_row
    for row in range(section_row + 1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is not None and str(ws.cell(row=row, column=1).value).strip().startswith("#"):
            boundary_row = row - 1; break
    def to_int(v):
        try: return int(v)
        except: return None
    skin_configs = [
        {"id": int(str(hero_id) + "001"), "name": f"{hero_name}-工笔白描", "suffix": "_xiangao"},
        {"id": int(str(hero_id) + "002"), "name": f"{hero_name}-动态原画", "suffix": ""},
        {"id": int(str(hero_id) + "003"), "name": f"{hero_name}-兼工带写", "suffix": "_gangman"},
    ]
    print(f"\n[Step 3 - Item: 武将皮肤道具]")
    for skin in skin_configs:
        skin_id = skin["id"]
        ref_row = section_row
        for row in range(section_row + 1, boundary_row + 1):
            val = to_int(ws.cell(row=row, column=1).value)
            if val is not None and val < skin_id:
                ref_row = row
        new_row = insert_row_below(ws, ref_row, max_col)
        boundary_row += 1
        icon = f"UI/Images/Item/ui_s1_daoju_wujiang_{hero_pinyin_lower}{skin['suffix']}.png"
        desc = f"{hero_name}的武将皮肤，获得后可以在游戏中改变武将{hero_name}的卡牌形象"
        ws.cell(row=new_row, column=1).value = skin_id
        ws.cell(row=new_row, column=2).value = skin["name"]
        ws.cell(row=new_row, column=3).value = "HeroSkin"
        ws.cell(row=new_row, column=4).value = 3
        ws.cell(row=new_row, column=6).value = 1
        ws.cell(row=new_row, column=7).value = 1
        ws.cell(row=new_row, column=8).value = 0
        ws.cell(row=new_row, column=11).value = 0
        ws.cell(row=new_row, column=12).value = hero_item_id
        ws.cell(row=new_row, column=22).value = icon
        ws.cell(row=new_row, column=24).value = 1
        ws.cell(row=new_row, column=25).value = 1
        ws.cell(row=new_row, column=26).value = 0
        ws.cell(row=new_row, column=27).value = 1
        ws.cell(row=new_row, column=28).value = desc
        ws.cell(row=new_row, column=37).value = 1
        print(f"  Row {new_row}: id={skin_id} name={skin['name']}")
    wb.save(item_path); wb.close()
    print(f"  OK Item.xlsx (3 skins)")

# ---- STEP 4 ----
def step4_item_avatar(base_dir, hero_name, hero_id, hero_pinyin):
    """Insert hero avatar item in #武将形象道具 section of Item.xlsx."""
    import re
    item_path = os.path.join(base_dir, "Item.xlsx")
    wb = openpyxl.load_workbook(item_path)
    ws = wb["道具表|Item"]
    max_col = ws.max_column

    # Calculate avatar ID: 102 + last 4 digits of hero_id
    hero_id_str = str(hero_id)
    avatar_id = int("102" + hero_id_str[-4:])

    # Find #武将形象道具 section
    section_row = None
    for row in range(5, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is not None and str(ws.cell(row=row, column=1).value).strip() == "#武将形象道具":
            section_row = row; break

    if section_row is None:
        print("  WARNING: #武将形象道具 not found, skipping"); wb.close(); return

    # Find boundary (next section marker)
    boundary_row = ws.max_row
    for row in range(section_row + 1, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1).value
        if c1 is not None and str(c1).strip().startswith("#"):
            boundary_row = row - 1; break

    def to_int(v):
        try: return int(v)
        except: return None

    # Find insertion position by ID order, also remember template row
    ref_row = section_row
    template_row = None
    for row in range(section_row + 1, boundary_row + 1):
        val = to_int(ws.cell(row=row, column=1).value)
        if val is not None:
            template_row = row
            if val < avatar_id:
                ref_row = row

    if template_row is None:
        print("  WARNING: No existing rows in #武将形象道具 section, cannot get template"); wb.close(); return

    new_row = insert_row_below(ws, ref_row, max_col)
    print(f"\n[Step 4 - Item: 武将形象道具] Ref={ref_row} -> New={new_row}")

    # Extract template values
    template_icon = ws.cell(row=template_row, column=22).value or ""
    template_desc = ws.cell(row=template_row, column=28).value or ""
    template_name = ws.cell(row=template_row, column=2).value or ""

    # Extract template hero name from col B format "{hero}-形象"
    template_hero = str(template_name).replace("-形象", "") if template_name else ""

    # Replace pinyin in icon path (last segment before .png → new hero_pinyin PascalCase)
    icon_match = re.search(r'_(.*?)\.png$', str(template_icon))
    if icon_match:
        old_pinyin = icon_match.group(1)
        new_icon = str(template_icon).replace(old_pinyin, str(hero_pinyin), 1)
    else:
        new_icon = template_icon

    # Replace hero name in description
    if template_hero and template_hero != hero_name:
        new_desc = str(template_desc).replace(template_hero, hero_name)
    else:
        new_desc = template_desc

    # Override specific fields (all others already copied from ref_row)
    ws.cell(row=new_row, column=1).value = avatar_id
    ws.cell(row=new_row, column=2).value = f"{hero_name}-形象"
    ws.cell(row=new_row, column=4).value = 3
    ws.cell(row=new_row, column=22).value = new_icon
    ws.cell(row=new_row, column=28).value = new_desc

    print(f"  id={avatar_id} name={hero_name}-形象")
    wb.save(item_path); wb.close()
    print(f"  OK Item.xlsx (avatar item)")

# ---- STEP 5 ----
def step5_spine(base_dir, hero_name, hero_id):
    """Insert dynamic portrait spine in #动态皮肤 basic section (after marker)."""
    sp_path = os.path.join(base_dir, "HeroSkinSpine_英雄皮肤Spine.xlsx")
    wb = openpyxl.load_workbook(sp_path)
    ws = wb["英雄皮肤Spine|HeroSkinSpine"]
    max_col = ws.max_column

    # Find #动态皮肤 marker
    marker_row = None
    for row in range(5, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is not None and str(ws.cell(row=row, column=1).value).strip() == "#动态皮肤":
            marker_row = row; break
    if marker_row is None:
        print("  WARNING: #动态皮肤 not found, skipping spine"); wb.close(); return

    def to_int(v):
        try: return int(v)
        except: return None

    dynamic_skin_id = int(str(hero_id) + "002")

    # Find insertion point AFTER marker: last row with col2 containing '动态皮肤' and skin_id < new_id
    ref_row = marker_row  # default: right after marker
    for row in range(marker_row + 1, ws.max_row + 1):
        c2 = ws.cell(row=row, column=2).value
        if c2 is not None and '动态皮肤' in str(c2):
            val = to_int(ws.cell(row=row, column=1).value)
            if val is not None and val < dynamic_skin_id:
                ref_row = row
            else:
                break

    new_row = insert_row_below(ws, ref_row, max_col)
    print(f"\n[Step 5 - HeroSkinSpine: #动态皮肤] Ref={ref_row} -> New={new_row}")
    ws.cell(row=new_row, column=1).value = dynamic_skin_id
    ws.cell(row=new_row, column=2).value = f"{hero_name}动态皮肤"
    ws.cell(row=new_row, column=3).value = True
    ws.cell(row=new_row, column=10).value = True
    ws.cell(row=new_row, column=12).value = True
    print(f"  id={dynamic_skin_id} C3/C10/C12=True")
    wb.save(sp_path); wb.close()
    print(f"  OK HeroSkinSpine")# ---- STEP 6 ----
def step6_heroskinitem(base_dir, hero_name, hero_id, hero_pinyin_lower):
    """Insert 3 skin detail rows in #武将皮肤 section, ordered by hero_id."""
    hs_path = os.path.join(base_dir, "HeroSkinItem_英雄皮肤.xlsx")
    wb = openpyxl.load_workbook(hs_path)
    ws = wb["英雄皮肤|HeroSkinItem"]
    max_col = ws.max_column

    def to_int(v):
        try: return int(v)
        except: return None

    # Find #武将皮肤 section marker
    section_row = None
    for row in range(5, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1).value
        if c1 is not None and str(c1).strip() == "#武将皮肤":
            section_row = row; break

    if section_row is None:
        print("  WARNING: #武将皮肤 section not found, skipping"); return wb, None, None

    # Find end of section (next section marker)
    end_row = ws.max_row
    for row in range(section_row + 1, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1).value
        if c1 is not None and str(c1).strip().startswith("#"):
            end_row = row - 1; break

    # Within #武将皮肤 section, find last row whose col2(hero_id) < new_hero_id
    ref_row = section_row
    for row in range(section_row + 1, end_row + 1):
        c2 = to_int(ws.cell(row=row, column=2).value)
        c1 = ws.cell(row=row, column=1).value
        if c2 is not None and c1 is not None:
            if c2 < hero_id:
                ref_row = row
            else:
                break

    # Go to last row of that hero's skin group
    last_hid = ws.cell(row=ref_row, column=2).value
    if last_hid is not None:
        for row in range(ref_row, end_row + 1):
            c2 = ws.cell(row=row, column=2).value
            if c2 is not None and str(c2) == str(last_hid):
                ref_row = row
            else:
                break

    skin1_id = int(str(hero_id) + "001")
    skin2_id = int(str(hero_id) + "002")
    skin3_id = int(str(hero_id) + "003")

    skin_rows = [
        {"sid": skin1_id, "c3": f"{hero_name}线稿皮肤", "raily": 1,
         "stype": "SkinLineSkin", "sname": "工笔白描",
         "pinyin": f"{hero_pinyin_lower}_xiangao",
         "get_way": "六合时邕解锁一定阶数获得", "collection": "ECollitionType_GBBM"},
        {"sid": skin2_id, "c3": f"{hero_name}动态皮肤", "raily": None,
         "stype": "SkinNormalDynamicsSkin", "sname": "动态原画",
         "pinyin": hero_pinyin_lower,
         "get_way": "六合时邕解锁一定阶数获得", "collection": None},
        {"sid": skin3_id, "c3": f"{hero_name}港漫皮肤", "raily": 2,
         "stype": "SkinHKComicsSkin", "sname": "兼工带写",
         "pinyin": f"{hero_pinyin_lower}_gangman",
         "get_way": "六合时邕解锁一定阶数获得", "collection": "ECollitionType_JGDX"},
    ]

    print(f"\n[Step 6 - HeroSkinItem]")
    for skin in skin_rows:
        new_row = insert_row_below(ws, ref_row, max_col)
        ref_row = new_row  # next insertion below this one
        ws.cell(row=new_row, column=1).value = skin["sid"]
        ws.cell(row=new_row, column=2).value = hero_id
        ws.cell(row=new_row, column=3).value = skin["c3"]
        if skin["raily"] is not None:
            ws.cell(row=new_row, column=4).value = skin["raily"]
        else:
            ws.cell(row=new_row, column=4).value = None
        ws.cell(row=new_row, column=5).value = skin["stype"]
        ws.cell(row=new_row, column=6).value = skin["sname"]
        ws.cell(row=new_row, column=7).value = skin.get("get_way", "待配置") # G(7)
        ws.cell(row=new_row, column=8).value = skin["pinyin"]
        ws.cell(row=new_row, column=16).value = "待配置"
        ws.cell(row=new_row, column=17).value = "待配置"
        ws.cell(row=new_row, column=18).value = skin.get("collection") # R(18)
        ws.cell(row=new_row, column=19).value = 1
        print(f"  Row {new_row}: id={skin['sid']} c3={skin['c3']} type={skin['stype']}")

    wb.save(hs_path)
    # Keep wb open for step 7 (line filling)
    return wb, ws, ref_row


# ---- STEP 7 ----
def step7_hero_lines(base_dir, hero_name, wb_hs, ws_hs, ref_row_hs):
    hl_path = os.path.join(base_dir, "HeroLines_武将台词表.xlsx")
    print(f"\n[Step 7 - HeroLines]")
    if not os.path.exists(hl_path):
        print("  WARNING: HeroLines not found, lines left empty"); wb_hs.save(ws_hs.parent.title if hasattr(ws_hs,'parent') else "");
        if wb_hs: wb_hs.close()
        return

    wb_hl = openpyxl.load_workbook(hl_path)
    ws_hl = wb_hl["武将台词|HeroLines"]
    line_data = find_hero_lines(ws_hl, hero_name)
    wb_hl.close()

    if not line_data["found"]:
        print(f"  WARNING: No lines found for hero '{hero_name}' in HeroLines")
        print(f"  Lines left empty - please configure manually")
        wb_hs.save(os.path.join(os.path.dirname(hl_path), "HeroSkinItem_英雄皮肤.xlsx"))
        wb_hs.close()
        return

    print(f"  all_lines:   {line_data['all_lines']}")
    print(f"  debut_lines: {line_data['debut_lines']}")
    print(f"  kill_lines:  {line_data['kill_lines']}")
    print(f"  dead_lines:  {line_data['dead_lines']}")
    print(f"  opt_lines:   {line_data['opt_lines']}")

    # The 3 skin rows are the last 3 rows we inserted (ref_row_hs, ref_row_hs-1, ref_row_hs-2)
    # But since we inserted sequentially, they are ref_row_hs-2, ref_row_hs-1, ref_row_hs
    for offset in [2, 1, 0]:
        target_row = ref_row_hs - offset
        if target_row < 5: continue
        set_text_cell(ws_hs, target_row, 11, line_data["all_lines"])
        set_text_cell(ws_hs, target_row, 12, line_data["debut_lines"])
        set_text_cell(ws_hs, target_row, 13, line_data["kill_lines"])
        set_text_cell(ws_hs, target_row, 14, line_data["dead_lines"])
        set_text_cell(ws_hs, target_row, 15, line_data["opt_lines"])
        print(f"  Filled lines in row {target_row}")

    hs_path = os.path.join(os.path.dirname(hl_path), "HeroSkinItem_英雄皮肤.xlsx")
    wb_hs.save(hs_path)
    wb_hs.close()
    print(f"  OK HeroSkinItem (lines filled)")


# ---- STEP 8 ----
def step8_heroui(base_dir, hero_name, hero_id, expansion_pack):
    ui_path = os.path.join(base_dir, "HeroUI_武将表现表.xlsx")
    wb = openpyxl.load_workbook(ui_path)
    ws = wb["武将表现配置|HeroUI"]
    max_col = ws.max_column

    def to_int(v):
        try: return int(v)
        except: return None
    # Find insertion point by hero_id order (col1)
    ref_row = 4
    for row in range(5, ws.max_row + 1):
        val = to_int(ws.cell(row=row, column=1).value)
        if val is not None and val < hero_id: ref_row = row
        else: break

    new_row = insert_row_below(ws, ref_row, max_col)
    print(f"\n[Step 8 - HeroUI] Ref={ref_row} -> New={new_row}")

    ws.cell(row=new_row, column=1).value = hero_id
    ws.cell(row=new_row, column=2).value = hero_name
    ws.cell(row=new_row, column=3).value = None  # 声音类型
    ws.cell(row=new_row, column=4).value = expansion_pack
    ws.cell(row=new_row, column=5).value = None  # 武将获得描述
    ws.cell(row=new_row, column=6).value = None  # 简介(长)
    ws.cell(row=new_row, column=7).value = None  # 简介(短)
    ws.cell(row=new_row, column=8).value = None  # 考据
    ws.cell(row=new_row, column=9).value = None  # 评价
    ws.cell(row=new_row, column=10).value = 0    # 是否新武将
    ws.cell(row=new_row, column=11).value = "椰椰"
    ws.cell(row=new_row, column=12).value = "待配置"
    ws.cell(row=new_row, column=13).value = None  # 获得方式
    ws.cell(row=new_row, column=14).value = None  # 武将定位
    ws.cell(row=new_row, column=15).value = None  # 武将专属牌
    ws.cell(row=new_row, column=16).value = None  # 技能ID
    ws.cell(row=new_row, column=17).value = None  # 2v2胜率
    ws.cell(row=new_row, column=18).value = None  # 胜率显示优先级

    print(f"  hero_id={hero_id} name={hero_name} expansion={expansion_pack}")
    wb.save(ui_path); wb.close()
    print(f"  OK HeroUI")


# ---- MAIN ----
def main():
    parser = argparse.ArgumentParser(description="Configure a new hero item")
    parser.add_argument("base_dir", help="Config directory")
    parser.add_argument("hero_name", help="Hero name")
    parser.add_argument("quality", choices=["传说", "史诗", "稀有", "普通"], help="Hero quality")
    args = parser.parse_args()

    quality_info = QUALITY_MAP[args.quality]
    print("=" * 60)
    print(f" 配置武将道具: {args.hero_name} ({args.quality})")
    print("=" * 60)

    # Step 1
    hero_id, hero_pinyin, hero_pinyin_lower, hero_item_id, expansion_pack = \
        step1_read_hero(args.base_dir, args.hero_name)

    # Step 2
    step2_item_hero(args.base_dir, args.hero_name, hero_id, hero_item_id,
                    hero_pinyin_lower, quality_info)

    # Step 3
    step3_item_skins(args.base_dir, args.hero_name, hero_id, hero_item_id,
                     hero_pinyin_lower)

    # Step 4
    step4_item_avatar(args.base_dir, args.hero_name, hero_id, hero_pinyin)

    # Step 5
    step5_spine(args.base_dir, args.hero_name, hero_id)

    # Step 6
    wb_hs, ws_hs, ref_row_hs = step6_heroskinitem(
        args.base_dir, args.hero_name, hero_id, hero_pinyin_lower)

    # Step 7
    step7_hero_lines(args.base_dir, args.hero_name, wb_hs, ws_hs, ref_row_hs)

    # Step 8
    step8_heroui(args.base_dir, args.hero_name, hero_id, expansion_pack)

    print("")
    print("=" * 60)
    print(" 配置完成!")
    print(f"   武将: {args.hero_name}")
    print(f"   HeroID: {hero_id}")
    print(f"   HeroItemId: {hero_item_id}")
    print(f"   品质: {args.quality}")
    print(f"   皮肤ID: {hero_id}001, {hero_id}002, {hero_id}003")
    print("=" * 60)


if __name__ == "__main__":
    main()