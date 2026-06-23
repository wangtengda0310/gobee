#!/usr/bin/env python3
"""configure_hero_mengjiang.py - 配置武将萌将道具 (Item.xlsx #武将形象道具 区域)

引用 common-rules.md 通用规则:
  - 规则 1: 插入行 = 全量复制 + 覆盖差异
  - 规则 2: 列定位 = 动态表头映射
  - 规则 4: Skill 更新 = MD + 脚本双同步
"""

import argparse
import os
import sys
from copy import copy
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    print("错误: 需要 openpyxl 库"); sys.exit(1)


# ============== 公共工具函数 (来自 common-rules.md) ==============

def copy_cell_style(src_cell, dst_cell):
    """复制单元格的值+全部样式"""
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def copy_row_style(ws, src_row, dst_row, max_col):
    """从 src_row 全量复制值+样式到 dst_row"""
    for col in range(1, max_col + 1):
        copy_cell_style(ws.cell(row=src_row, column=col),
                        ws.cell(row=dst_row, column=col))


def safe_insert_rows(ws, start_row, count=1):
    """安全地在 start_row 位置插入 count 行，完整保留所有格式。"""
    from openpyxl.worksheet.dimensions import RowDimension

    max_row_before = ws.max_row
    max_col = ws.max_column

    # 从下往上逐行搬移数据+样式（倒序，避免覆盖）
    for row in range(max_row_before, start_row - 1, -1):
        new_row = row + count
        for col in range(1, max_col + 1):
            src = ws.cell(row=row, column=col)
            dst = ws.cell(row=new_row, column=col)
            dst.value = src.value
            if src.has_style:
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format
                dst.protection = copy(src.protection)

    # 清空插入区域
    for row in range(start_row, start_row + count):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None

    # 偏移合并单元格引用
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= start_row:
            mc.shift(0, count)

    # 偏移行尺寸
    for r in range(max_row_before, start_row - 1, -1):
        rd_old = ws.row_dimensions.get(r)
        if rd_old is not None:
            rd_new = ws.row_dimensions.get(r + count)
            if rd_new is None:
                rd_new = RowDimension(worksheet=ws, index=r + count)
                ws.row_dimensions[r + count] = rd_new
            rd_new.height = rd_old.height
            rd_new.hidden = rd_old.hidden
            try:
                rd_new.customFormat = rd_old.customFormat
            except AttributeError:
                pass
            try:
                rd_new.customHeight = rd_old.customHeight
            except AttributeError:
                pass

    return list(range(start_row, start_row + count))


# ============== 业务逻辑 ==============

def step1_read_hero_pinyin(base_dir, hero_name):
    """Step 1: 读取 Hero.xlsx，获取武将 C 列 PascalCase 拼音"""
    hero_path = os.path.join(base_dir, "Hero.xlsx")
    wb = openpyxl.load_workbook(hero_path)
    ws = wb["武将表|Hero"]

    hero_pinyin = None
    for row in range(5, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        if cell_val is not None and str(cell_val).strip() == hero_name:
            hero_pinyin = ws.cell(row=row, column=3).value
            break

    wb.close()
    if hero_pinyin is None:
        print(f"ERROR: 在 Hero.xlsx 中未找到武将 '{hero_name}'")
        sys.exit(1)

    pinyin_lower = str(hero_pinyin).lower()
    print(f"\n[Step 1 - Hero.xlsx] {hero_name}: Pinyin(PascalCase)={hero_pinyin}, Pinyin(lower)={pinyin_lower}")
    return hero_pinyin, pinyin_lower


def step2_configure_mengjiang(base_dir, hero_name, pinyin_lower):
    """Step 2: 在 Item.xlsx #武将形象道具 区域配置萌将道具"""
    import re
    item_path = os.path.join(base_dir, "Item.xlsx")
    wb = openpyxl.load_workbook(item_path)
    ws = wb["道具表|Item"]
    max_col = ws.max_column

    # ---- 定位 #武将形象道具 区域 ----
    section_row = None
    for row in range(5, ws.max_row + 1):
        if ws.cell(row=row, column=1).value is not None \
           and str(ws.cell(row=row, column=1).value).strip() == "#武将形象道具":
            section_row = row
            break

    if section_row is None:
        print("ERROR: 未找到 #武将形象道具 区域标记")
        wb.close()
        sys.exit(1)

    # 找区域边界
    boundary_row = ws.max_row
    for row in range(section_row + 1, ws.max_row + 1):
        c1 = ws.cell(row=row, column=1).value
        if c1 is not None and str(c1).strip().startswith("#"):
            boundary_row = row - 1
            break

    # ---- 找到最后一个萌将道具行 ----
    def to_int(v):
        try: return int(v)
        except: return None

    last_mj_row = None
    max_mj_id = 0
    for row in range(section_row + 1, boundary_row + 1):
        name_val = str(ws.cell(row=row, column=2).value or "")
        if "·萌将" in name_val:
            last_mj_row = row
            rid = to_int(ws.cell(row=row, column=1).value)
            if rid is not None and rid > max_mj_id:
                max_mj_id = rid

    if last_mj_row is None:
        print("ERROR: 在 #武将形象道具 区域未找到任何萌将道具")
        wb.close()
        sys.exit(1)

    new_id = max_mj_id + 1
    print(f"\n[Step 2 - Item: 萌将道具]")
    print(f" 最后萌将: Row {last_mj_row}, ID={max_mj_id}, Name={ws.cell(row=last_mj_row, column=2).value}")
    print(f" 新萌将ID={new_id}")

    # ---- 在最后一个萌将行下方插入新行 (规则 1: 全量复制 + 覆盖差异) ----
    safe_insert_rows(ws, last_mj_row + 1, 1)
    new_row = last_mj_row + 1
    copy_row_style(ws, last_mj_row, new_row, max_col)

    # ---- 提取模板行的 icon 和 desc，做替换 ----
    template_icon = str(ws.cell(row=last_mj_row, column=22).value or "")

    # Icon: 替换拼音部分 (最后一段 _xxx_Qban.png 中的 pinyin)
    icon_match = re.search(r'_([a-z]+)_Qban\.png$', template_icon, re.IGNORECASE)
    if icon_match:
        old_pinyin = icon_match.group(1)
        new_icon = template_icon.replace(f"_{old_pinyin}_Qban.png", f"_{pinyin_lower}_Qban.png")
    else:
        # fallback: 构造标准路径
        new_icon = f"UI/Images/Item/ui_s1_daoju_pifuxingxiang_{pinyin_lower}_Qban.png"

    # Desc: 格式固定
    new_desc = f"{hero_name}·萌将形象，可以在形象系统中使用，将自己的形象设置为{hero_name}·萌将形象"

    # ---- 只覆盖差异字段 (规则 1) ----
    ws.cell(row=new_row, column=1).value = new_id
    ws.cell(row=new_row, column=2).value = f"{hero_name}·萌将"
    ws.cell(row=new_row, column=4).value = 4  # 品质
    ws.cell(row=new_row, column=14).value = "{1000002;1000}"  # 分解道具
    ws.cell(row=new_row, column=22).value = new_icon
    ws.cell(row=new_row, column=28).value = new_desc

    wb.save(item_path)
    wb.close()

    print(f" Row {new_row}: id={new_id}, name={hero_name}·萌将")
    print(f" icon: {new_icon}")
    print(f" desc: {new_desc}")
    print(f"\n✅ 萌将道具配置完成!")
    return new_row, new_id


# ============== Main ==============

def main():
    parser = argparse.ArgumentParser(description="配置名将杀萌将道具")
    parser.add_argument("base_dir", help="配置文件目录 (如 /home/.../名将杀配置/)")
    parser.add_argument("hero_name", help="武将名称 (如 孙权)")
    args = parser.parse_args()

    print("=" * 60)
    print(f" 配置萌将道具: {args.hero_name}·萌将")
    print("=" * 60)

    # Step 1: 读取 Hero.xlsx 获取拼音
    hero_pinyin, pinyin_lower = step1_read_hero_pinyin(args.base_dir, args.hero_name)

    # Step 2: 配置萌将道具
    new_row, new_id = step2_configure_mengjiang(args.base_dir, args.hero_name, pinyin_lower)

    print(f"\n{'='*60}")
    print(f" 配置完成!")
    print(f"   武将: {args.hero_name}")
    print(f"   萌将道具ID: {new_id}")
    print(f"   道具名称: {args.hero_name}·萌将")
    print(f"   插入位置: Item.xlsx Row {new_row}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
